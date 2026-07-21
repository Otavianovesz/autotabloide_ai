"""
Ponte universal — exportar/importar o acervo em Excel (R-118 da Fase 11)
=======================================================================
O acervo vai para ``.xlsx`` e volta. A disciplina é a MESMA da portabilidade
(Bloco D): casar por **CHAVE NATURAL** (nome_sanitizado + marca, I1), NUNCA por
id; conflito NUNCA em silêncio (I2); prévia→confirma (nada grava na análise).

Diferenças para o ``.atpkg``: a planilha é uma ponte de DADOS, não de arquivos
— não leva foto nem caminho de máquina (I3); leva nome/marca/categoria/EAN/
preço/sabor/peso/validade/flags. O roundtrip export→(edita)→import é idêntico
por chave natural.

Fluxo em duas fases (a UI mostra o relatório entre elas), igual ao pacote:

    analise = analisar_planilha("acervo.xlsx")   # nada é gravado
    ...humano decide os conflitos...
    rel = aplicar_importacao_planilha(analise, decisoes)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from sqlalchemy import select

from app.core.paths import SystemRoot
from app.core.portabilidade import Decisao, _decimal, _norm, chave_natural

# ordem das colunas da planilha (o cabeçalho)
COLUNAS = [
    "Nome", "Marca", "Categoria", "EAN", "Preço", "Sabor",
    "Peso", "Unidade", "Validade", "Bebida alcoólica", "Selo +18",
    "Marca própria",
]

_VERDADE = {"sim", "s", "1", "true", "verdadeiro", "x"}


def _root(raiz) -> SystemRoot:
    if isinstance(raiz, SystemRoot):
        return raiz
    return SystemRoot(raiz).criar_estrutura() if raiz else SystemRoot().criar_estrutura()


def _flag_texto(v: bool) -> str:
    return "Sim" if v else ""


def _flag_bool(v) -> bool:
    return str(v or "").strip().lower() in _VERDADE


def _preco_br(valor) -> str:
    return f"{valor:.2f}".replace(".", ",") if valor is not None else ""


def _peso_texto(valor) -> str:
    if valor is None:
        return ""
    try:
        return f"{Decimal(valor).normalize():f}".replace(".", ",")
    except (InvalidOperation, ValueError):
        return str(valor)


def _data_texto(d) -> str:
    return d.strftime("%d/%m/%Y") if isinstance(d, (date, datetime)) else ""


def _parse_data(txt) -> date | None:
    # o openpyxl (data_only) devolve datetime para célula de DATA real — aceita
    # o objeto direto, não só o texto (senão a validade digitada no Excel some).
    if isinstance(txt, datetime):
        return txt.date()
    if isinstance(txt, date):
        return txt
    txt = str(txt or "").strip()
    if not txt:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y",
                "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    return None


def _preco_planilha(txt) -> Decimal | None:
    """Aceita '5,90' (BR), '5.90' (US) e número — o formato do operador."""
    if txt is None or str(txt).strip() == "":
        return None
    s = str(txt).strip().replace("R$", "").replace(" ", "")
    if s.count(",") and s.count("."):        # "1.234,56" → milhar + decimal
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


# --- exportar ----------------------------------------------------------------------

def _linha_do_produto(p, nome_categoria: str | None) -> dict:
    return {
        "Nome": p.nome_sanitizado or "",
        "Marca": p.marca or "",
        "Categoria": nome_categoria or "",
        "EAN": p.ean or "",
        "Preço": _preco_br(p.preco_atual),
        "Sabor": p.sabor or "",
        "Peso": _peso_texto(p.peso_valor),
        "Unidade": p.peso_unidade or "",
        "Validade": _data_texto(p.validade_item),
        "Bebida alcoólica": _flag_texto(bool(p.bebida_alcoolica)),
        "Selo +18": _flag_texto(bool(p.selo_mais18)),
        "Marca própria": _flag_texto(bool(p.marca_propria)),
    }


def exportar_acervo_xlsx(destino: str | Path, raiz=None) -> Path:
    """Grava o acervo vivo (não excluído) numa planilha .xlsx. Sem foto (I3)."""
    from openpyxl import Workbook

    from app.core.database import Database
    from app.core.models import Categoria, Produto

    destino = Path(destino)
    if destino.suffix.lower() != ".xlsx":
        destino = destino.with_suffix(".xlsx")
    root = _root(raiz)

    wb = Workbook()
    ws = wb.active
    ws.title = "Acervo"
    ws.append(COLUNAS)
    db = Database(root).init()
    try:
        with db.Session() as s:
            cats = {c.id: c.nome for c in s.execute(select(Categoria)).scalars()}
            for p in s.execute(select(Produto).where(
                    Produto.excluido_em.is_(None)).order_by(
                    Produto.nome_sanitizado)).scalars():
                linha = _linha_do_produto(p, cats.get(p.categoria_id))
                ws.append([linha[c] for c in COLUNAS])
    finally:
        db.engine.dispose()
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))
    return destino


# --- analisar (fase 1: nada é gravado) ---------------------------------------------

@dataclass
class ConflitoPlanilha:
    """Mesma chave natural, dados divergentes — o humano decide (I2)."""

    id_decisao: str                       # "produto:<nome>|<marca>"
    rotulo: str
    campos: list[str]
    id_local: int = 0
    plano: dict = field(default_factory=dict)     # plano completo da planilha (aplicar)
    linha: dict = field(default_factory=dict)     # valores da planilha (exibição)
    local: dict = field(default_factory=dict)     # valores atuais (exibição)


@dataclass
class AnalisePlanilha:
    caminho: str = ""
    novos: list[dict] = field(default_factory=list)      # linhas só da planilha
    identicos: list[str] = field(default_factory=list)   # nomes já iguais
    conflitos: list[ConflitoPlanilha] = field(default_factory=list)
    ignoradas: list[str] = field(default_factory=list)   # lixo (sem nome)
    avisos: list[str] = field(default_factory=list)
    _raiz: str | None = None


def _ler_linhas(caminho: Path) -> list[dict]:
    """Lê a planilha em dicts {coluna: valor}, tolerando cabeçalho reordenado.

    Ignora linhas totalmente vazias; usa o cabeçalho da 1ª linha para mapear as
    colunas (aceita ausência de colunas — o que faltar vira '')."""
    from openpyxl import load_workbook

    wb = load_workbook(str(caminho), read_only=True, data_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not linhas:
        return []
    cabecalho = [str(c).strip() if c is not None else "" for c in linhas[0]]
    idx = {nome: i for i, nome in enumerate(cabecalho)}
    out: list[dict] = []
    for bruto in linhas[1:]:
        if bruto is None or all(c is None or str(c).strip() == "" for c in bruto):
            continue                                     # linha em branco: pula
        d = {}
        for col in COLUNAS:
            i = idx.get(col)
            v = bruto[i] if (i is not None and i < len(bruto)) else None
            if v is None:
                d[col] = ""
            elif isinstance(v, (date, datetime)):
                # célula de DATA real do Excel: normaliza p/ dd/mm/aaaa (não
                # "2026-07-06 00:00:00", que o _parse_data não casava → validade
                # sumia calada e o roundtrip virava conflito fantasma)
                d[col] = _data_texto(v)
            else:
                d[col] = str(v).strip()
        out.append(d)
    return out


def _plano_da_linha(linha: dict) -> dict:
    """Linha da planilha → plano comparável (mesma forma do produto local)."""
    return {
        "nome": linha.get("Nome", "").strip(),
        "marca": linha.get("Marca", "").strip(),
        "categoria": linha.get("Categoria", "").strip(),
        "ean": linha.get("EAN", "").strip(),
        "preco": _preco_planilha(linha.get("Preço")),
        "sabor": linha.get("Sabor", "").strip(),
        "peso": _preco_planilha(linha.get("Peso")),
        "unidade": linha.get("Unidade", "").strip(),
        "validade": _parse_data(linha.get("Validade")),
        "alcool": _flag_bool(linha.get("Bebida alcoólica")),
        "mais18": _flag_bool(linha.get("Selo +18")),
        "marca_propria": _flag_bool(linha.get("Marca própria")),
    }


def _plano_local(p, nome_categoria: str | None) -> dict:
    return {
        "id": p.id,
        "nome": p.nome_sanitizado or "",
        "marca": p.marca or "",
        "categoria": nome_categoria or "",
        "ean": p.ean or "",
        "preco": p.preco_atual,
        "sabor": p.sabor or "",
        "peso": p.peso_valor,
        "unidade": p.peso_unidade or "",
        "validade": p.validade_item,
        "alcool": bool(p.bebida_alcoolica),
        "mais18": bool(p.selo_mais18),
        "marca_propria": bool(p.marca_propria),
    }


_FLAG_CHAVE = {"bebida alcoólica": "alcool", "selo +18": "mais18",
               "marca própria": "marca_propria"}


def _exibir_campos(plano: dict, campos: list[str]) -> dict:
    """Valores formatados SÓ dos campos que divergem — o preview do conflito
    mostra o antes/depois do que realmente mudou (I2), não só o preço."""
    m: dict[str, str] = {}
    for c in campos:
        if c == "preço":
            m[c] = _preco_br(plano.get("preco")) or "—"
        elif c == "categoria":
            m[c] = plano.get("categoria") or "—"
        elif c == "EAN":
            m[c] = plano.get("ean") or "—"
        elif c == "sabor":
            m[c] = plano.get("sabor") or "—"
        elif c == "peso":
            m[c] = _peso_texto(plano.get("peso")) or "—"
        elif c == "validade":
            m[c] = _data_texto(plano.get("validade")) or "—"
        elif c in _FLAG_CHAVE:
            m[c] = "Sim" if plano.get(_FLAG_CHAVE[c]) else "Não"
    return m


def _campos_divergentes(local: dict, linha: dict) -> list[str]:
    """O que diverge entre o produto local e a linha da planilha (sem foto)."""
    difere: list[str] = []
    if (local["preco"] or None) != (linha["preco"] or None):
        difere.append("preço")
    if _norm(local["categoria"]) != _norm(linha["categoria"]):
        difere.append("categoria")
    if _norm(local["ean"]) != _norm(linha["ean"]):
        difere.append("EAN")
    if _norm(local["sabor"]) != _norm(linha["sabor"]):
        difere.append("sabor")
    lp = (local["peso"] if local["peso"] is None else Decimal(local["peso"]),
          _norm(local["unidade"]))
    ln = (linha["peso"], _norm(linha["unidade"]))
    if lp != ln:
        difere.append("peso")
    if local["validade"] != linha["validade"]:
        difere.append("validade")
    for campo, nome in (("alcool", "bebida alcoólica"), ("mais18", "selo +18"),
                        ("marca_propria", "marca própria")):
        if local[campo] != linha[campo]:
            difere.append(nome)
    return difere


def analisar_planilha(caminho: str | Path, raiz=None) -> AnalisePlanilha:
    """Fase 1: lê a planilha e compara com o acervo por chave natural. Nada grava."""
    from app.core.database import Database
    from app.core.models import Categoria, Produto

    caminho = Path(caminho)
    if not caminho.is_file():
        raise FileNotFoundError(f"planilha não encontrada: {caminho}")
    root = _root(raiz)
    analise = AnalisePlanilha(caminho=str(caminho), _raiz=str(root.raiz))

    linhas = _ler_linhas(caminho)
    db = Database(root).init()
    try:
        with db.Session() as s:
            cats = {c.id: c.nome for c in s.execute(select(Categoria)).scalars()}
            locais = {chave_natural(p.nome_sanitizado, p.marca):
                      _plano_local(p, cats.get(p.categoria_id))
                      for p in s.execute(select(Produto).where(
                          Produto.excluido_em.is_(None))).scalars()}
    finally:
        db.engine.dispose()

    vistos: set[tuple] = set()
    for linha in linhas:
        plano = _plano_da_linha(linha)
        if not plano["nome"]:
            analise.ignoradas.append("linha sem nome ignorada")
            continue
        chave = chave_natural(plano["nome"], plano["marca"])
        if chave in vistos:                       # planilha com dup interna
            analise.avisos.append(
                f"“{plano['nome']}” aparece mais de uma vez na planilha — "
                "só a primeira conta")
            continue
        vistos.add(chave)
        local = locais.get(chave)
        if local is None:
            analise.novos.append(plano)
            continue
        campos = _campos_divergentes(local, plano)
        if not campos:
            analise.identicos.append(plano["nome"])
            continue
        analise.conflitos.append(ConflitoPlanilha(
            id_decisao=f"produto:{chave[0]}|{chave[1]}",
            rotulo=plano["nome"] + (f" ({plano['marca']})" if plano["marca"] else ""),
            campos=campos, id_local=local["id"], plano=plano,
            linha=_exibir_campos(plano, campos),     # antes/depois do que mudou
            local=_exibir_campos(local, campos)))
    return analise


# --- aplicar (fase 2: grava exigindo decisão para todo conflito) -------------------

@dataclass
class RelatorioPlanilha:
    produtos_novos: list[str] = field(default_factory=list)
    conflitos_resolvidos: list[tuple[str, str]] = field(default_factory=list)
    ignoradas: int = 0
    avisos: list[str] = field(default_factory=list)

    def resumo(self) -> str:
        partes = []
        if self.produtos_novos:
            partes.append(f"{len(self.produtos_novos)} produtos novos")
        if self.conflitos_resolvidos:
            partes.append(f"{len(self.conflitos_resolvidos)} conflitos resolvidos")
        if self.ignoradas:
            partes.append(f"{self.ignoradas} linhas ignoradas")
        return "Importação da planilha concluída: " + (", ".join(partes) or "nada a fazer") + "."


def aplicar_importacao_planilha(analise: AnalisePlanilha,
                                decisoes: dict[str, Decisao] | None = None,
                                raiz=None) -> RelatorioPlanilha:
    """Fase 2: grava. Todo conflito EXIGE decisão explícita (nada em silêncio, I2).

    Produto novo nasce por chave natural; conflito com USAR_PACOTE atualiza os
    campos da planilha; MANTER_LOCAL não toca; MANTER_AMBOS cria variante."""
    from app.core.modo import exigir_escrita
    exigir_escrita()                     # R-131: mil produtos pela planilha
    #                                      é tão escrita quanto editar um
    from app.core.database import Database
    from app.core.models import Categoria, Produto

    decisoes = decisoes or {}
    pendentes = [c.rotulo for c in analise.conflitos
                 if c.id_decisao not in decisoes]
    if pendentes:
        raise ValueError(
            "conflitos sem decisão (I2 — nada se resolve em silêncio): "
            + "; ".join(pendentes))

    root = _root(raiz if raiz is not None else analise._raiz)
    rel = RelatorioPlanilha(ignoradas=len(analise.ignoradas))
    db = Database(root).init()
    try:
        with db.Session() as s:
            def _categoria_id(nome_cat: str | None):
                if not nome_cat:
                    return None
                row = next((c for c in s.execute(select(Categoria)).scalars()
                            if _norm(c.nome) == _norm(nome_cat)), None)
                if row is None:
                    row = Categoria(nome=nome_cat)
                    s.add(row)
                    s.flush()
                return row.id

            locais = {chave_natural(p.nome_sanitizado, p.marca): p
                      for p in s.execute(select(Produto).where(
                          Produto.excluido_em.is_(None))).scalars()}

            def _aplicar_campos(prod: Produto, plano: dict) -> None:
                prod.categoria_id = _categoria_id(plano["categoria"])
                prod.preco_atual = plano["preco"]
                prod.ean = plano["ean"] or None
                prod.sabor = plano["sabor"] or None
                prod.peso_valor = plano["peso"]
                prod.peso_unidade = plano["unidade"] or None
                prod.validade_item = plano["validade"]
                prod.bebida_alcoolica = plano["alcool"]
                prod.selo_mais18 = plano["mais18"]
                prod.marca_propria = plano["marca_propria"]

            # novos — id novo, casando por chave natural (E-A2: revalida)
            for plano in analise.novos:
                if chave_natural(plano["nome"], plano["marca"]) in locais:
                    rel.avisos.append(
                        f"“{plano['nome']}” já existe agora — pulado (o acervo "
                        "mudou desde a análise)")
                    continue
                prod = Produto(nome_bruto=plano["nome"],
                               nome_sanitizado=plano["nome"], marca=plano["marca"] or None)
                _aplicar_campos(prod, plano)
                s.add(prod)
                s.flush()
                locais[chave_natural(plano["nome"], plano["marca"])] = prod
                rel.produtos_novos.append(plano["nome"])

            # conflitos — só com decisão
            for c in analise.conflitos:
                decisao = decisoes[c.id_decisao]
                if decisao is Decisao.MANTER_LOCAL:
                    rel.conflitos_resolvidos.append((c.rotulo, decisao.value))
                    continue
                # E-A2 (espelha portabilidade): casa por CHAVE NATURAL, NUNCA por
                # id — se o produto foi renomeado entre analisar e aplicar, a chave
                # sumiu e NÃO se grava no produto errado (I1); avisa (I2).
                chave = chave_natural(c.plano["nome"], c.plano["marca"])
                prod = locais.get(chave)
                if decisao is Decisao.USAR_PACOTE and prod is None:
                    rel.avisos.append(
                        f"“{c.rotulo}” mudou de identidade (renomeado?) desde a "
                        "análise — pulado; re-analise a planilha")
                    continue
                if decisao is Decisao.USAR_PACOTE:
                    _aplicar_campos(prod, c.plano)
                elif decisao is Decisao.MANTER_AMBOS:
                    plano = c.plano
                    variante = Produto(
                        nome_bruto=plano["nome"],
                        nome_sanitizado=_nome_variante(s, plano["nome"], plano["marca"]),
                        marca=plano["marca"] or None)
                    _aplicar_campos(variante, plano)
                    s.add(variante)
                    s.flush()
                # MANTER_LOCAL: nada
                rel.conflitos_resolvidos.append((c.rotulo, decisao.value))
            s.commit()
    finally:
        db.engine.dispose()
    return rel


def _nome_variante(session, nome: str, marca: str | None) -> str:
    from app.core.models import Produto
    existentes = {chave_natural(p.nome_sanitizado, p.marca)
                  for p in session.execute(select(Produto).where(
                      Produto.excluido_em.is_(None))).scalars()}
    candidato, n = f"{nome} (planilha)", 2
    while chave_natural(candidato, marca) in existentes:
        candidato = f"{nome} (planilha {n})"
        n += 1
    return candidato
