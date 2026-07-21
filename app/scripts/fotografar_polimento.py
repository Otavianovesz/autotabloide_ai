"""Galeria NATIVA do passe de POLIMENTO pré-F12.

Fotografa cada tela/diálogo tocado pelo passe: a fusão de duplicatas lado a
lado, o editor de ajuste de imagem, o comparador de versões, a foto do acervo,
as opções do relâmpago (QR), a inteligência com eixos, o import de Excel em
PT-BR, a barra da Fábrica com RG-53 (larga E estreita com "···"), o modo
planilha com colar/massa e as manchetes do papel de texto.

SEM offscreen de propósito (o offscreen renderiza glifos como caixas); janelas
com WA_DontShowOnScreen — nada pisca. Rodar::

    python -m app.scripts.fotografar_polimento saida_polimento/claro
    python -m app.scripts.fotografar_polimento saida_polimento/escuro --tema=escuro
"""

from __future__ import annotations

import os
import sys
import tempfile
from datetime import datetime
from decimal import Decimal
from pathlib import Path


def _processar(n: int = 3) -> None:
    from PySide6.QtWidgets import QApplication
    for _ in range(n):
        QApplication.processEvents()


def _grab(widget, pasta: Path, nome: str) -> None:
    _processar()
    pasta.mkdir(parents=True, exist_ok=True)
    widget.grab().save(str(pasta / nome))
    print(f"  {nome}")


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    pasta = Path(args[0] if args else "saida_polimento/claro")
    tema = next((a.split("=", 1)[1] for a in sys.argv[1:]
                 if a.startswith("--tema=")), None)

    # raiz TEMPORÁRIA semeada (nunca o acervo vivo)
    sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "tests"))
    import seeds_portabilidade as seeds
    tmp = Path(tempfile.mkdtemp(prefix="polimento_"))
    root = seeds.raiz(tmp, "raiz")
    os.environ["AUTOTABLOIDE_ROOT"] = str(root.raiz)
    seeds.add_produto(root, "Arroz Tio João 5kg", "Tio João", "24.90",
                      foto=seeds.png("#C0392B"), categoria="Mercearia")
    seeds.add_produto(root, "Arroz Tio João 5kg", "Tio João", "19.90")
    seeds.add_produto(root, "Café Pilão 500g", "Pilão", "9.90",
                      foto=seeds.png("#2C3E50"), categoria="Mercearia")
    seeds.add_produto(root, "Detergente Ypê 500ml", "Ypê", "1.99",
                      foto=seeds.png("#27AE60"), categoria="Limpeza")

    from PySide6.QtCore import Qt
    from PySide6.QtWidgets import QApplication

    from app.qt.design.tema import aplicar_tema
    app = QApplication.instance() or QApplication([])
    aplicar_tema(app, tema) if tema else aplicar_tema(app)
    from app.qt.design.polimento import instalar_polimento
    instalar_polimento(app)
    DONT = Qt.WidgetAttribute.WA_DontShowOnScreen

    from app.qt.telas import servico

    # --- 1) fusão de duplicatas (lado a lado) --------------------------------
    from app.qt.telas.duplicatas_dialog import DuplicatasDialog
    pares = servico.pares_duplicatas()
    dlg = DuplicatasDialog(pares)
    dlg.setAttribute(DONT, True)
    dlg.show()
    _grab(dlg, pasta, "duplicatas_lado_a_lado.png")
    dlg.close()

    # --- 2) ajuste de imagem (girar/cortar com prévia) -----------------------
    from PIL import Image, ImageDraw
    foto = tmp / "produto.png"
    img = Image.new("RGB", (420, 320), (245, 245, 247))
    d = ImageDraw.Draw(img)
    d.rounded_rectangle([150, 40, 270, 290], radius=18, fill=(192, 57, 43))
    d.rectangle([170, 120, 250, 200], fill=(250, 250, 250))
    img = img.rotate(-8, expand=True, fillcolor=(245, 245, 247))
    img.save(foto)
    from app.qt.telas.ajuste_imagem_dialog import AjusteImagemDialog
    dlg = AjusteImagemDialog(str(foto))
    dlg.setAttribute(DONT, True)
    dlg.show()
    _grab(dlg, pasta, "ajuste_imagem.png")
    dlg.close()

    # --- 3) foto do acervo (picker) ------------------------------------------
    from app.qt.telas.acervo_picker_dialog import AcervoPickerDialog
    dlg = AcervoPickerDialog()
    dlg.setAttribute(DONT, True)
    dlg.show()
    _grab(dlg, pasta, "acervo_picker.png")
    dlg.close()

    # --- 4) comparador de versões --------------------------------------------
    from app.core.paths import SystemRoot
    from app.images.biblioteca import BibliotecaImagens
    from app.qt.telas.almoxarifado import HistoricoImagensDialog
    bib = BibliotecaImagens(SystemRoot().biblioteca_imagens)
    pid = seeds.produto_por_chave(root, "Café Pilão 500g", "Pilão")["id"]
    nova = tmp / "nova.png"
    nova.write_bytes(seeds.png("#8E44AD"))
    bib.ingerir(pid, str(nova))
    dlg = HistoricoImagensDialog(pid)
    dlg.setAttribute(DONT, True)
    dlg.show()
    dlg.lista.setCurrentRow(0)
    _grab(dlg, pasta, "historico_comparador.png")
    dlg.close()

    # --- 5) opções do relâmpago (QR) -----------------------------------------
    from app.qt.telas.relampago_dialog import RelampagoDialog
    dlg = RelampagoDialog("Café Pilão 500g", kit=True)
    dlg.com_qr.setChecked(True)
    dlg.qr_texto.setText("https://belobrasil.com.br/encarte")
    dlg.setAttribute(DONT, True)
    dlg.show()
    _grab(dlg, pasta, "relampago_opcoes.png")
    dlg.close()

    # --- 6) inteligência (eixos rotulados + estados vazios com craft) --------
    from app.core import projetos as _proj
    from app.qt.telas.servico import ItemMesa

    def _ed(evento, quando, pares_nome_preco):
        return {"evento": evento, "criado_em": quando,
                "itens": [ItemMesa(n, p, "VERDE", n, produto_id=i + 1).to_dict()
                          for i, (n, p) in enumerate(pares_nome_preco)]}

    _originais = _proj.historico_edicoes
    _proj.historico_edicoes = lambda limite=None: [
        _ed("Quintou", datetime(2026, 5, 7, 10),
            [("Arroz Tio João 5kg", "26,90"), ("Café Pilão 500g", "10,90")]),
        _ed("Quintou", datetime(2026, 6, 4, 10),
            [("Arroz Tio João 5kg", "23,50")]),
        _ed("FDS", datetime(2026, 6, 20, 10),
            [("Arroz Tio João 5kg", "24,90"), ("Café Pilão 500g", "9,90")]),
        _ed("Quintou", datetime(2026, 7, 16, 10),
            [("Arroz Tio João 5kg", "19,90")]),
    ]
    try:
        from app.qt.telas.inteligencia_dialog import InteligenciaDialog
        dlg = InteligenciaDialog()
        dlg.setAttribute(DONT, True)
        dlg.show()
        _grab(dlg, pasta, "inteligencia_saude.png")
        abas = dlg.findChildren.__self__            # (o QTabWidget é o único)
        from PySide6.QtWidgets import QTabWidget
        tabs = dlg.findChild(QTabWidget)
        tabs.setCurrentIndex(2)                     # histórico (eixos)
        _grab(dlg, pasta, "inteligencia_historico_eixos.png")
        dlg.close()
    finally:
        _proj.historico_edicoes = _originais

    # --- 7) importar Excel em PT-BR (com conflito) ---------------------------
    from openpyxl import load_workbook

    from app.core import excel_acervo as X
    xlsx = X.exportar_acervo_xlsx(tmp / "acervo.xlsx", raiz=root)
    wb = load_workbook(str(xlsx))
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "Café Pilão 500g":
            row[4].value = "8,49"
    wb.save(str(xlsx))
    from app.qt.telas.importar_planilha_dialog import ImportarPlanilhaDialog
    dlg = ImportarPlanilhaDialog(X.analisar_planilha(xlsx, raiz=root))
    dlg.setAttribute(DONT, True)
    dlg.show()
    _grab(dlg, pasta, "importar_planilha_ptbr.png")
    dlg.close()

    # --- 8) barra da Fábrica: RG-53 larga e estreita -------------------------
    from app.qt.telas.fabrica import FabricaTela
    from app.qt.telas.servico import ItemMesa as _IM
    f = FabricaTela()
    f.setAttribute(DONT, True)
    f._itens = [_IM("Café Pilão 500g", "8,49", "VERDE", "Café Pilão 500g",
                    preco_de="9,90", categoria="Mercearia",
                    imagem=str(root.biblioteca_imagens / f"{pid}/atual.png"))]
    f._recarregar_lista()
    f.show()
    f.resize(1720, 760)
    f._reflow_barra()
    _processar()
    f.lista.setCurrentRow(0)
    _grab(f, pasta, "fabrica_larga.png")
    f.resize(1000, 720)
    f._barra_fabrica.resize(1000, 44)
    f._reflow_barra()
    _grab(f._barra_fabrica, pasta, "fabrica_barra_estreita_720p.png")
    f.close()

    # --- 9) modo planilha (dica de colar/massa) ------------------------------
    from app.qt.telas.mesa import MesaTela
    from app.qt.telas.planilha_dialog import DialogoPlanilha
    m = MesaTela()
    m.setAttribute(DONT, True)
    m.show()
    m._itens = [
        _IM("Arroz Tio João 5kg", "24,90", "VERDE", "Arroz Tio João 5kg",
            categoria="Mercearia"),
        _IM("Café Pilão 500g", "9,90", "VERDE", "Café Pilão 500g"),
        _IM("Detergente Ypê 500ml", "2x 5,00", "VERDE", "Detergente Ypê 500ml"),
    ]
    dlg = DialogoPlanilha(m, m)
    dlg.setAttribute(DONT, True)
    dlg.show()
    _grab(dlg, pasta, "planilha_colar_massa.png")
    dlg.close()
    m.close()

    # --- 10) papel de texto com manchetes (degradadas, sem IA) ---------------
    from app.qt.design.papel_texto_ui import _dialogo_cls
    from app.rendering.model import PapelTexto
    dlg = _dialogo_cls()(None, papel=PapelTexto.LIVRE,
                         contexto={"evento": "Quintou"})
    from app.ai.enriquecimento import sugerir_manchetes
    dlg.combo_manchetes.clear()
    dlg.combo_manchetes.addItem("Manchetes sugeridas…")
    for mch in sugerir_manchetes("Quintou", None):
        dlg.combo_manchetes.addItem(mch)
    dlg._atualizar_visibilidade()
    dlg.setAttribute(DONT, True)
    dlg.show()
    _grab(dlg, pasta, "papel_manchetes.png")
    dlg.close()

    print(f"galeria em {pasta}")
    # teardown nativo do Qt no Windows derruba o processo DEPOIS dos arquivos
    from app.qt.workers import encerrar_todos
    encerrar_todos(espera_ms=1000)
    app.closeAllWindows()
    _processar()
    os._exit(0)


if __name__ == "__main__":
    raise SystemExit(main())
