"""
FASE 11 — Cartaz completo, impressão e inteligência (testes)
============================================================
Prova por CONTEÚDO (mm, pixels, bytes) — nunca por "não deu exceção":

  * a biblioteca de layouts de cartaz sai no tamanho físico exato (mm);
  * o % de desconto é CALCULADO (de−por)/de e realmente DESENHA no cartaz;
  * a cartaz-relâmpago vai do produto ao PDF em mm, com pré-voo (I2);
  * o kit ponta-de-gôndola é coerente (cartaz + etiquetas do MESMO dado);
  * o QR opcional é gerado localmente e é não-destrutivo.

Os blocos B–F adicionam a este arquivo (impressão, Excel, inteligência,
adversarial). A régua de mm é ``mediabox`` do pypdf (pontos = mm/25.4*72).
"""

from __future__ import annotations

import tempfile
from decimal import Decimal
from pathlib import Path

import pytest
from PIL import Image
from pypdf import PdfReader

from app.rendering import cartaz
from app.rendering.compositor import (
    DadosProduto,
    _rect_px,
    compor_pagina,
    percentual_desconto,
    texto_composto_legal,
)
from app.rendering.model import PapelTexto, Regiao, Retangulo, TipoRegiao
from app.qt.telas import servico

MM_POR_PT = 25.4 / 72.0


def _pdf_mm(caminho, pagina=0) -> tuple[float, float]:
    box = PdfReader(str(caminho)).pages[pagina].mediabox
    return (float(box.width) * MM_POR_PT, float(box.height) * MM_POR_PT)


def _regiao_por_papel(layout, papel: PapelTexto) -> Regiao:
    for pag in layout.paginas:
        for slot in pag.slots:
            for reg in slot.regioes:
                if getattr(reg, "papel_texto", None) == papel:
                    return reg
    raise AssertionError(f"nenhuma região com papel {papel}")


def _recorte(img: Image.Image, reg: Regiao, dpi: int) -> Image.Image:
    x, y, w, h = _rect_px(reg.rect, dpi)
    return img.crop((x, y, x + w, y + h))


def _tem_tinta(recorte: Image.Image, fundo=(255, 255, 255), tol=30) -> int:
    """Quantos pixels destoam do fundo branco — mede se algo foi DESENHADO."""
    dados = recorte.convert("RGB").tobytes()
    n = 0
    for i in range(0, len(dados), 3):
        r, g, b = dados[i], dados[i + 1], dados[i + 2]
        if abs(r - fundo[0]) > tol or abs(g - fundo[1]) > tol or abs(b - fundo[2]) > tol:
            n += 1
    return n


def _foto(tmp: Path, cor="orange") -> str:
    p = tmp / "foto.png"
    Image.new("RGB", (400, 400), cor).save(p)
    return str(p)


# --- R-109: % de desconto CALCULADO ------------------------------------------------

def test_percentual_desconto_calculado():
    assert percentual_desconto(Decimal("12.99"), Decimal("9.99")) == 23
    assert percentual_desconto(Decimal("10.00"), Decimal("5.00")) == 50
    assert percentual_desconto(Decimal("100"), Decimal("75")) == 25
    # arredondamento claro (HALF_UP)
    assert percentual_desconto(Decimal("8.00"), Decimal("6.60")) == 18  # 17.5→18


def test_percentual_desconto_some_sem_de_ou_sem_desconto():
    assert percentual_desconto(None, Decimal("9.99")) is None       # sem "de"
    assert percentual_desconto(Decimal("9"), None) is None          # sem "por"
    assert percentual_desconto(Decimal("9"), Decimal("10")) is None  # de<por
    assert percentual_desconto(Decimal("9"), Decimal("9")) is None   # iguais
    assert percentual_desconto(Decimal("0"), Decimal("0")) is None   # de<=0


def test_desconto_region_desenha_e_some_por_conteudo():
    """O selo -XX% aparece no cartaz quando há desconto e SOME sem "de"
    (prova por pixels: o recorte da região tem tinta, ou não tem)."""
    lay = cartaz.layout_cartaz_a5()
    reg = _regiao_por_papel(lay, PapelTexto.DESCONTO)

    com = DadosProduto("Arroz", preco_por=Decimal("9.90"), preco_de=Decimal("14.90"))
    img_com = compor_pagina(lay, lay.paginas[0], com)
    assert _tem_tinta(_recorte(img_com, reg, lay.dpi)) > 50   # desenhou "-34%"

    sem = DadosProduto("Arroz", preco_por=Decimal("9.90"), preco_de=None)
    img_sem = compor_pagina(lay, lay.paginas[0], sem)
    assert _tem_tinta(_recorte(img_sem, reg, lay.dpi)) == 0    # nada sem "de"

    # e o texto exato bate com o cálculo
    assert texto_composto_legal(reg, com) == "-34%"
    assert texto_composto_legal(reg, sem) == ""


# --- R-105: biblioteca de layouts no tamanho físico exato (mm) ---------------------

@pytest.mark.parametrize("nome,esperado", [
    ("Cartaz 10×15 — exemplo", (100.0, 150.0)),
    ("A4 retrato (210×297 mm)", (210.0, 297.0)),
    ("A4 paisagem (297×210 mm)", (297.0, 210.0)),
    ("Meia folha A5 (148×210 mm)", (148.0, 210.0)),
    ("Etiqueta de prateleira (100×70 mm)", (100.0, 70.0)),
])
def test_biblioteca_layouts_mm(nome, esperado):
    lay = cartaz.PRESETS_CARTAZ[nome]()
    assert (round(lay.largura_mm, 1), round(lay.altura_mm, 1)) == esperado
    assert lay.dpi == 300
    # 1 item por página, 1 slot (decisão travada)
    assert len(lay.paginas) == 1 and len(lay.paginas[0].slots) == 1


def test_todo_layout_tem_a_cadeia_de_regioes():
    """Cada layout de cartaz tem NOME, preço "por", desconto e validade — a
    mesma cadeia produto→slot; a etiqueta dispensa só a IMAGEM."""
    for nome, fn in cartaz.PRESETS_CARTAZ.items():
        lay = fn()
        papeis = {getattr(r, "papel_texto", None)
                  for s in lay.paginas[0].slots for r in s.regioes}
        tipos = {r.tipo for s in lay.paginas[0].slots for r in s.regioes}
        assert PapelTexto.DESCONTO in papeis, nome
        assert PapelTexto.VALIDADE in papeis, nome
        assert TipoRegiao.NOME in tipos and TipoRegiao.PRECO in tipos, nome


# --- R-110: cartaz-relâmpago (produto → PDF em mm) ---------------------------------

def test_cartaz_relampago_mm_e_desconto():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        prod = {"nome": "Café Pilão 500g", "preco": "9,99", "preco_de": "12,99",
                "imagem": _foto(tmp), "validade": "24/07", "categoria": "Mercearia"}
        saida, avisos = servico.cartaz_relampago(prod, str(tmp / "c.pdf"))
        assert Path(saida).exists()
        larg, alt = _pdf_mm(saida)
        assert abs(larg - 100.0) < 1.0 and abs(alt - 150.0) < 1.0
        assert avisos == []          # tudo presente: sem pendência
        # o % calculado bate: (12,99−9,99)/12,99 = 23%
        d = servico.dados_cartaz_de_produto(prod)
        assert percentual_desconto(d.preco_de, d.preco_por) == 23


def test_cartaz_relampago_prevoo_avisa_mas_exporta():
    """I2: sem foto/de/validade AVISA — mas o PDF sai (nunca bloqueia)."""
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        prod = {"nome": "Sabão X", "preco": "2,99", "imagem": None,
                "validade": "", "categoria": ""}
        saida, avisos = servico.cartaz_relampago(prod, str(tmp / "c.pdf"))
        assert Path(saida).exists()          # exportou mesmo assim
        texto = " · ".join(avisos).lower()
        assert "sem foto" in texto
        assert "de" in texto                 # sem preço "de"
        assert "validade" in texto           # RG-58: validade nunca calada


def test_cartaz_relampago_carimba_rascunho():
    """Decisão travada: relâmpago é sempre RASCUNHO (sem projeto aprovado)."""
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        prod = {"nome": "Item", "preco": "5,00", "preco_de": "9,00",
                "imagem": _foto(tmp), "validade": "30/07"}
        lay = cartaz.layout_cartaz_exemplo()
        dados = servico.dados_cartaz_de_produto(prod)
        limpo = compor_pagina(lay, lay.paginas[0], dados)
        saida, _ = servico.cartaz_relampago(prod, str(tmp / "c.pdf"), layout=lay)
        # a página do PDF NÃO é a composição limpa — a marca d'água entrou
        pagina = PdfReader(str(saida)).pages[0]
        # comparar por bytes da imagem embutida vs a limpa exportada à parte
        from app.rendering.export import exportar_pdf
        exportar_pdf(limpo, str(tmp / "limpo.pdf"), lay.dpi)
        a = pagina.images[0].data
        b = PdfReader(str(tmp / "limpo.pdf")).pages[0].images[0].data
        assert a != b               # rascunho carimbou (conteúdo diferente)


# --- R-113: kit ponta-de-gôndola coerente ------------------------------------------

def test_kit_paginas_tamanhos_e_coerencia():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        prod = {"nome": "Detergente 500ml", "preco": "1,99", "preco_de": "3,49",
                "imagem": _foto(tmp), "validade": "31/07"}
        saida, avisos = servico.gerar_kit_gondola(
            prod, str(tmp / "kit.pdf"), n_etiquetas=2)
        r = PdfReader(str(saida))
        assert len(r.pages) == 3                       # cartaz + 2 etiquetas
        assert abs(_pdf_mm(saida, 0)[0] - 148.0) < 1.0  # A5
        assert abs(_pdf_mm(saida, 1)[0] - 100.0) < 1.0  # etiqueta
        assert abs(_pdf_mm(saida, 2)[1] - 70.0) < 1.0
        # coerência: TODAS as páginas saem do MESMO dado — a página do cartaz
        # é byte-idêntica à composição manual com o mesmo DadosProduto
        dados = servico.dados_cartaz_de_produto(prod)
        lay = cartaz.layout_cartaz_a5()
        esperado, _ = servico._compor_cartaz(lay, dados, rascunho=True)
        from app.rendering.export import exportar_pdf
        exportar_pdf(esperado, str(tmp / "esp.pdf"), lay.dpi)
        assert (r.pages[0].images[0].data
                == PdfReader(str(tmp / "esp.pdf")).pages[0].images[0].data)


def test_kit_coerencia_reflete_o_preco_por_conteudo():
    """Se o preço do produto muda, o cartaz e a etiqueta mudam JUNTOS
    (prova de que o preço flui de uma fonte só — muta e vê a diferença)."""
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        base = {"nome": "Item", "preco": "5,00", "preco_de": "9,00",
                "imagem": _foto(tmp), "validade": "30/07"}
        lay_e = cartaz.layout_etiqueta()
        et_barato, _ = servico._compor_cartaz(
            lay_e, servico.dados_cartaz_de_produto(base), rascunho=False)
        caro = dict(base, preco="8,00")
        et_caro, _ = servico._compor_cartaz(
            lay_e, servico.dados_cartaz_de_produto(caro), rascunho=False)
        assert et_barato.tobytes() != et_caro.tobytes()   # o preço mudou o desenho


# --- R-114: QR opcional, local, não-destrutivo -------------------------------------

def test_qr_gera_local_e_muda_por_texto():
    from app.rendering.qr import gerar_qr
    a, av_a = gerar_qr("https://belobrasil.com/encarte", 240)
    b, av_b = gerar_qr("https://outro.com/x", 240)
    assert a is not None and b is not None and av_a is None and av_b is None
    assert a.size == (240, 240)
    assert a.tobytes() != b.tobytes()             # texto diferente = QR diferente
    vazio, av = gerar_qr("", 240)
    assert vazio is None and av                    # texto vazio avisa


def test_qr_aplicar_nao_destrutivo():
    from app.rendering.qr import aplicar_qr
    base = Image.new("RGB", (400, 600), "white")
    fora, aviso = aplicar_qr(base, "https://x.com", lado_px=80)
    assert aviso is None
    assert base.tobytes() != fora.tobytes()        # a cópia mudou
    assert base.getpixel((10, 10)) == (255, 255, 255)  # a original intacta


# ============================================================================
# BLOCO B — impressão (R-106 2-em-1, R-108 lote, R-112 imprimir direto)
# ============================================================================

from app.rendering.units import mm_para_px  # noqa: E402


def _a5(cor, dpi=300):
    return Image.new("RGB", (round(mm_para_px(148, dpi)),
                             round(mm_para_px(210, dpi))), cor)


def test_impor_2em1_medido_e_por_conteudo():
    """R-106/passo 63: A4 paisagem (297×210) com 2 A5, cada metade no lugar
    certo (esquerda vermelha, direita azul) — medido em px pelo DPI."""
    from app.rendering.imposicao import impor_2em1
    dpi = 300
    verm, azul = _a5((220, 30, 30), dpi), _a5((30, 30, 220), dpi)
    folhas = impor_2em1([verm, azul], dpi)
    assert len(folhas) == 1
    f = folhas[0]
    assert abs(f.width * 25.4 / dpi - 297.0) < 0.6    # A4 paisagem
    assert abs(f.height * 25.4 / dpi - 210.0) < 0.6
    meia = f.width // 2
    assert f.getpixel((meia // 2, f.height // 2)) == (220, 30, 30)   # esquerda
    assert f.getpixel((meia + meia // 2, f.height // 2)) == (30, 30, 220)  # direita


def test_impor_2em1_impar_e_marcas_de_corte():
    from app.rendering.imposicao import impor_2em1
    dpi = 300
    a = _a5((10, 200, 10), dpi)
    # 3 cartazes → 2 folhas (a última com só a metade esquerda)
    assert len(impor_2em1([a, a, a], dpi)) == 2
    # as marcas de corte mudam a folha (conteúdo, não só "não deu exceção")
    sem = impor_2em1([a, a], dpi, marcas_corte=False)[0]
    com = impor_2em1([a, a], dpi, marcas_corte=True)[0]
    assert sem.tobytes() != com.tobytes()


@pytest.mark.parametrize("fn_nome,esperado", [
    ("layout_cartaz_a4_retrato", (210.0, 297.0)),
    ("layout_cartaz_a4_paisagem", (297.0, 210.0)),
    ("layout_cartaz_a5", (148.0, 210.0)),
])
def test_impressao_respeita_mm_e_orientacao(fn_nome, esperado):
    """R-112/passo 33: a impressão direta respeita o tamanho físico e a
    orientação — um QPrinter em modo PDF grava a página que a régua de mm mede."""
    from PySide6.QtPrintSupport import QPrinter
    from PySide6.QtWidgets import QApplication

    from app.rendering.impressao import imprimir_imagens
    QApplication.instance() or QApplication([])
    lay = getattr(cartaz, fn_nome)()
    with tempfile.TemporaryDirectory() as td:
        img = compor_pagina(lay, lay.paginas[0],
                            DadosProduto("T", preco_por=Decimal("9.90"),
                                         preco_de=Decimal("12.90")))
        saida = str(Path(td) / "print.pdf")
        pr = QPrinter(QPrinter.PrinterMode.HighResolution)
        pr.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        pr.setOutputFileName(saida)
        n = imprimir_imagens([img, img], lay, pr)
        assert n == 2
        larg, alt = _pdf_mm(saida)
        assert abs(larg - esperado[0]) < 1.0
        assert abs(alt - esperado[1]) < 1.0
        assert len(PdfReader(saida).pages) == 2


def test_lote_por_categoria_filtra_certo():
    """R-108: o lote reusa filtrar_itens — só a categoria escolhida sai."""
    from app.qt.telas.servico import ItemMesa, filtrar_itens
    itens = [
        ItemMesa("A", "1,00", "VERDE", "Arroz", categoria="Mercearia"),
        ItemMesa("B", "2,00", "VERDE", "Picanha", categoria="Açougue"),
        ItemMesa("C", "3,00", "VERDE", "Feijão", categoria="Mercearia"),
    ]
    acougue = filtrar_itens(itens, categoria="Açougue")
    assert [it.nome for it in acougue] == ["Picanha"]
    mercearia = filtrar_itens(itens, categoria="Mercearia")
    assert {it.nome for it in mercearia} == {"Arroz", "Feijão"}


def test_fabrica_prontos_para_saida_respeita_lote():
    """A Fábrica filtra os cartazes prontos pela categoria do lote (R-108)."""
    from PySide6.QtWidgets import QApplication

    from app.qt.telas.fabrica import FabricaTela
    from app.qt.telas.servico import ItemMesa
    QApplication.instance() or QApplication([])
    tela = FabricaTela()
    tela._itens = [
        ItemMesa("Arroz", "9,90", "VERDE", "Arroz", preco_de="12,90",
                 categoria="Mercearia"),
        ItemMesa("Picanha", "39,90", "VERDE", "Picanha", preco_de="49,90",
                 categoria="Açougue"),
    ]
    tela._recarregar_lista()
    tela.combo_categoria.setCurrentText("Açougue")
    prontos = tela._prontos_para_saida()
    assert [it.nome for it in prontos] == ["Picanha"]
    tela.close()


# ============================================================================
# BLOCO C — inteligência SÓ LEITURA + Excel por chave natural
# ============================================================================

from datetime import date, datetime  # noqa: E402

from app.tests import seeds_portabilidade as seeds  # noqa: E402


def _ed(evento, quando, itens):
    from app.qt.telas.servico import ItemMesa
    return {"evento": evento, "criado_em": quando,
            "itens": [it.to_dict() if isinstance(it, ItemMesa) else it
                      for it in itens]}


def _it(nome, preco, *, pid=None, cat=None, foto=None):
    from app.qt.telas.servico import ItemMesa
    return ItemMesa(nome, preco, "VERDE", nome, produto_id=pid, categoria=cat,
                    imagem=foto)


# --- R-115/120/121/117/123: inteligência (pura, injetável) -------------------------

def _historico_exemplo():
    return [
        _ed("Quintou", datetime(2025, 7, 20, 10),
            [_it("Arroz", "24,90", pid=1), _it("Café", "9,90", pid=2)]),
        _ed("FDS", datetime(2026, 3, 1, 10),
            [_it("Arroz", "22,90", pid=1), _it("Feijão", "7,50", pid=3)]),
        _ed("Quintou", datetime(2026, 7, 18, 10),
            [_it("Arroz", "19,90", pid=1)]),
    ]


def test_historico_de_preco_por_chave_natural():
    from app.qt.telas import inteligencia as I
    s = I.serie_de_um(_historico_exemplo(), ("pid", 1))     # o Arroz
    assert [str(p.preco) for p in s["pontos"]] == ["24.90", "22.90", "19.90"]
    assert str(s["menor"]) == "19.90"                       # menor do ano marcado
    assert len(s["menor_marcado"]) == 1


def test_ranking_e_sazonal_por_identidade():
    from app.qt.telas import inteligencia as I
    edi = _historico_exemplo()
    rank = I.ranking_ofertados(edi)
    assert rank[0]["nome"] == "Arroz" and rank[0]["edicoes"] == 3
    # ano passado nesta semana (2026-07-19 → 2025-07-20 Quintou)
    saz = {x["nome"] for x in I.memoria_sazonal(edi, hoje=date(2026, 7, 19))}
    assert saz == {"Arroz", "Café"}


def test_relatorio_da_edicao():
    from app.qt.telas import inteligencia as I
    itens = [_it("Arroz", "22,90", cat="Mercearia"),
             _it("Picanha", "39,90", cat="Açougue", foto=None)]
    rel = I.relatorio_edicao(itens)
    assert rel["total"] == 2
    assert rel["por_categoria"] == {"Açougue": 1, "Mercearia": 1}
    assert rel["sem_foto"] == 2
    assert str(rel["preco_min"]) == "22.90" and str(rel["preco_max"]) == "39.90"


def test_divergencia_de_precos_dispara_por_identidade():
    from decimal import Decimal

    from app.qt.telas import inteligencia as I
    # o MESMO item (uid u1) em 2 páginas com preços diferentes → divergência;
    # itens distintos (u2) com o mesmo preço não disparam
    oc = [("u1", "Arroz", Decimal("19.90"), 1),
          ("u1", "Arroz", Decimal("22.90"), 2),
          ("u2", "Café", Decimal("9.90"), 1)]
    div = I.divergencia_de_precos(oc)
    assert len(div) == 1 and div[0]["identidade"] == "u1"
    assert set(div[0]["precos"]) == {"19.90", "22.90"}
    # coincidência de nome com preço igual NÃO é divergência
    assert I.divergencia_de_precos(
        [("u1", "X", Decimal("5"), 1), ("u2", "X", Decimal("5"), 2)]) == []


def test_divergencias_no_mapa_por_uid():
    """R-123 no estado da Mesa: o MESMO item (uid) em 2 slots com preços
    diferentes dispara; produtos distintos (uid distinto) não."""
    from decimal import Decimal

    from app.qt.telas import inteligencia as I
    from app.rendering.compositor import DadosProduto
    dados = {
        "s1": DadosProduto("Arroz", preco_por=Decimal("19.90")),
        "s2": DadosProduto("Arroz", preco_por=Decimal("22.90")),
        "s3": DadosProduto("Café", preco_por=Decimal("9.90")),
    }
    mapa = {"s1": "uidA", "s2": "uidA", "s3": "uidB"}       # s1,s2 = mesmo item
    div = I.divergencias_no_mapa(dados, mapa)
    assert len(div) == 1 and div[0]["identidade"] == "uidA"


# --- R-118: Excel export/import por chave natural (I1/I2/I3) ------------------------

def _semear_acervo(tmp_path):
    root = seeds.raiz(tmp_path, "loja")
    seeds.add_produto(root, "Arroz 5kg", "Camil", "24.90", categoria="Mercearia")
    seeds.add_produto(root, "Cerveja Amstel 269ml", "Amstel", "2.99",
                      categoria="Bebidas", mais18=True)
    seeds.add_produto(root, "Feijão 1kg", "Kicaldo", "8.50", categoria="Mercearia")
    return root


def test_excel_roundtrip_idempotente(tmp_path):
    """R-118/passo 66: export→import sem editar = acervo idêntico (0 conflitos,
    tudo idêntico por chave natural)."""
    from app.core import excel_acervo as X
    root = _semear_acervo(tmp_path)
    xlsx = X.exportar_acervo_xlsx(tmp_path / "acervo.xlsx", raiz=root)
    a = X.analisar_planilha(xlsx, raiz=root)
    assert a.novos == [] and a.conflitos == []
    assert len(a.identicos) == 3


def test_excel_export_sem_caminho_absoluto(tmp_path):
    """I3: a planilha é ponte de DADOS — sem foto, sem caminho de máquina."""
    from openpyxl import load_workbook

    from app.core import excel_acervo as X
    root = _semear_acervo(tmp_path)
    xlsx = X.exportar_acervo_xlsx(tmp_path / "acervo.xlsx", raiz=root)
    wb = load_workbook(str(xlsx))
    ws = wb.active
    cabecalho = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert not any("imagem" in str(c).lower() or "caminho" in str(c).lower()
                   for c in cabecalho)
    texto = "\n".join(str(c.value) for row in ws.iter_rows() for c in row)
    assert str(tmp_path) not in texto and ".png" not in texto


def test_excel_import_conflito_por_chave_natural(tmp_path):
    """R-118/passo 65: editar o preço na planilha vira conflito por chave
    natural; sem decisão barra (I2); com USAR_PACOTE atualiza o produto certo."""
    from openpyxl import load_workbook

    from app.core import excel_acervo as X
    from app.core.portabilidade import Decisao
    root = _semear_acervo(tmp_path)
    xlsx = X.exportar_acervo_xlsx(tmp_path / "acervo.xlsx", raiz=root)
    wb = load_workbook(str(xlsx))
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "Arroz 5kg":
            row[4].value = "19,90"                     # coluna Preço
    wb.save(str(xlsx))

    a = X.analisar_planilha(xlsx, raiz=root)
    assert len(a.conflitos) == 1 and "preço" in a.conflitos[0].campos
    with pytest.raises(ValueError):                    # I2: sem decisão barra
        X.aplicar_importacao_planilha(a, {}, raiz=root)
    dec = {c.id_decisao: Decisao.USAR_PACOTE for c in a.conflitos}
    X.aplicar_importacao_planilha(a, dec, raiz=root)
    assert seeds.produto_por_chave(root, "Arroz 5kg", "Camil")["preco"] == "19.90"
    # o OUTRO produto não foi tocado (a chave natural não confunde)
    assert seeds.produto_por_chave(root, "Feijão 1kg", "Kicaldo")["preco"] == "8.50"


def test_excel_import_ignora_lixo(tmp_path):
    """Passo 67: linha em branco / sem nome → ignorada com honestidade (I2)."""
    from openpyxl import load_workbook

    from app.core import excel_acervo as X
    root = _semear_acervo(tmp_path)
    xlsx = X.exportar_acervo_xlsx(tmp_path / "acervo.xlsx", raiz=root)
    wb = load_workbook(str(xlsx))
    ws = wb.active
    ws.append([])                                      # linha em branco
    ws.append(["", "", "Mercearia", "", "5,00"])       # sem nome
    wb.save(str(xlsx))
    a = X.analisar_planilha(xlsx, raiz=root)
    assert len(a.identicos) == 3                        # os 3 reais entram
    assert a.ignoradas                                  # o lixo é reportado


def test_excel_import_nomes_iguais_marcas_diferentes(tmp_path):
    """Adversarial (passo 65): mesmo nome, marca diferente = produtos DIFERENTES
    (a chave natural inclui a marca) — nunca troca em silêncio."""
    from app.core import excel_acervo as X
    root = seeds.raiz(tmp_path, "loja")
    seeds.add_produto(root, "Achocolatado", "Nescau", "8.00", categoria="Mercearia")
    xlsx = X.exportar_acervo_xlsx(tmp_path / "a.xlsx", raiz=root)
    # adiciona no Excel um "Achocolatado" de OUTRA marca
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx))
    ws = wb.active
    ws.append(["Achocolatado", "Toddy", "Mercearia", "", "7,00"])
    wb.save(str(xlsx))
    a = X.analisar_planilha(xlsx, raiz=root)
    assert len(a.novos) == 1 and a.novos[0]["marca"] == "Toddy"   # produto NOVO
    assert a.conflitos == []                                       # nunca conflito


# --- R-122/R-126: meta e saúde (DB) ------------------------------------------------

def test_meta_por_evento(tmp_path):
    from app.qt.telas import inteligencia as I
    root = seeds.raiz(tmp_path, "loja")
    I.definir_meta_evento("Quintou", 40, raiz=root)
    assert I.meta_evento("Quintou", raiz=root) == 40
    p = I.progresso_meta("Quintou", 32, raiz=root)
    assert p["texto"] == "32/40" and not p["atingiu"]
    assert I.progresso_meta("Quintou", 40, raiz=root)["atingiu"]


def test_saude_acervo(tmp_path):
    from app.qt.telas import inteligencia as I
    root = seeds.raiz(tmp_path, "loja")
    seeds.add_produto(root, "Arroz", "Camil", "24.90", foto=seeds.png("#0f0"),
                      categoria="Mercearia")
    seeds.add_produto(root, "Café", "Pilão", None, categoria="Mercearia")
    s = I.saude_acervo(raiz=root)
    assert s["total"] == 2
    assert s["com_foto"] == 1 and s["pct_foto"] == 50
    assert s["com_preco"] == 1 and s["com_categoria"] == 2


def test_inteligencia_e_so_leitura(tmp_path):
    """Passo 61: os relatórios NÃO alteram o acervo (contagem antes == depois)."""
    from app.qt.telas import inteligencia as I
    root = _semear_acervo(tmp_path)
    antes = seeds.contagens(root)["produtos"]
    I.saude_acervo(raiz=root)
    I.ranking_ofertados(_historico_exemplo())
    I.relatorio_edicao([_it("X", "1,00")])
    assert seeds.contagens(root)["produtos"] == antes    # nada mudou


# --- passo 60: VETOS ausentes (R-116/R-119/R-124/R-125) ----------------------------

def test_vetos_ausentes_no_codigo():
    """Passo 53/60/69/95: nenhum recurso vetado é IMPLEMENTADO — custo/margem de
    LUCRO (R-116), diário de alterações (R-119), backup em NUVEM (R-124), ERP
    (R-125). A varredura procura IDENTIFICADORES de código (a ausência da
    FEATURE), não a palavra em prosa — "nada sai para a nuvem" é tranquilizador,
    não um recurso; a própria lista de vetos numa docstring não conta."""
    import re
    from pathlib import Path as _P
    raiz = _P(__file__).resolve().parents[1]            # app/
    proibidos = re.compile(
        r"(margem_lucro|lucratividade|preco_custo|preco_de_custo|"
        r"custo_unitario|custo_produto|backup_nuvem|nuvem_sync|"
        r"sincroniz\w*_nuvem|cloud_sync|upload_nuvem|integra\w*_erp|"
        r"\berp_|diario_de_alteracoes|diario_alteracoes|log_de_alteracoes)",
        re.IGNORECASE)
    achados = []
    for py in raiz.rglob("*.py"):
        if "test" in py.name:
            continue
        for n, linha in enumerate(py.read_text(encoding="utf-8").splitlines(), 1):
            if proibidos.search(linha):
                achados.append(f"{py.relative_to(raiz)}:{n}: {linha.strip()}")
    assert not achados, "recurso vetado encontrado:\n" + "\n".join(achados)


# ============================================================================
# BLOCO D — disciplina: imposição, Excel e vetos
# ============================================================================

def test_2em1_nunca_no_tabloide():
    """Passo 64/81: o 2-em-1 é SÓ do cartaz (Fábrica). A Mesa (tabloide) não
    conhece imposição; só a Fábrica referencia `impor_2em1`. Guarda testada."""
    from pathlib import Path as _P
    app_dir = _P(__file__).resolve().parents[1]
    mesa = (app_dir / "qt" / "telas" / "mesa.py").read_text(encoding="utf-8")
    assert "impor_2em1" not in mesa and "imposicao" not in mesa
    assert "chk_2em1" not in mesa
    # entre as TELAS (qt/telas/), só a Fábrica (cartaz) chama impor_2em1 — o
    # tabloide (Mesa) nunca. (A galeria em scripts/ é demo, não é tela.)
    usa = [py.relative_to(app_dir).as_posix()
           for py in (app_dir / "qt" / "telas").rglob("*.py")
           if "test" not in py.name
           and "impor_2em1(" in py.read_text(encoding="utf-8")]
    assert usa == ["qt/telas/fabrica.py"], usa


def test_excel_sem_coluna_id(tmp_path):
    """Passo 65 (adversarial): a planilha NÃO carrega id — casar por id entre
    PCs é impossível por construção; só a chave natural liga (I1)."""
    from app.core import excel_acervo as X
    root = _semear_acervo(tmp_path)
    xlsx = X.exportar_acervo_xlsx(tmp_path / "a.xlsx", raiz=root)
    from openpyxl import load_workbook
    ws = load_workbook(str(xlsx)).active
    cabecalho = [str(c.value).strip().lower() for c in next(ws.iter_rows(max_row=1))]
    assert "id" not in cabecalho and "produto_id" not in cabecalho


def test_kit_prevoo_cartaz_true(tmp_path):
    """Passo 68: o kit respeita o pré-voo cartaz=True — sem foto/preço avisa
    ANTES do PDF (I2), mas ainda gera."""
    from app.core import excel_acervo  # noqa: F401  (garante import do pacote)
    prod = {"nome": "Item sem nada", "preco": "", "imagem": None, "validade": ""}
    saida, avisos = servico.gerar_kit_gondola(prod, str(tmp_path / "kit.pdf"))
    assert Path(saida).exists()
    texto = " · ".join(avisos).lower()
    assert "cartaz:" in texto and "etiqueta:" in texto     # os dois pré-voos
    assert "sem" in texto                                  # pendências visíveis


def test_intel_nao_grava_caminho_absoluto(tmp_path):
    """I3/passo 70: os relatórios da inteligência são dados em memória — não
    gravam caminho absoluto (nada de arquivo com caminho de máquina)."""
    from app.qt.telas import inteligencia as I
    rel = I.relatorio_edicao([_it("Arroz", "9,90", cat="Mercearia")])
    # o relatório é um dict puro; serializa sem nenhum caminho de disco
    import json
    texto = json.dumps(rel, default=str)
    assert str(tmp_path) not in texto and ":\\" not in texto


# ============================================================================
# BLOCO F — fechamento: jargão PT-BR, GIF
# ============================================================================

def test_jargao_pt_br_presente():
    """Passo 93: o jargão do dono aparece na UI — "cartaz-relâmpago", "dois por
    folha", "histórico de preço". Nada de termo técnico em inglês nos botões."""
    from pathlib import Path as _P
    app_dir = _P(__file__).resolve().parents[1]
    fab = (app_dir / "qt" / "telas" / "fabrica.py").read_text(encoding="utf-8")
    alm = (app_dir / "qt" / "telas" / "almoxarifado.py").read_text(encoding="utf-8")
    intel = (app_dir / "qt" / "telas" / "inteligencia_dialog.py").read_text(encoding="utf-8")
    assert "Dois por folha" in fab
    assert "Cartaz-relâmpago" in alm and "Kit ponta-de-gôndola" in alm
    assert "Histórico de preço" in intel or "Histórico de preço" in \
        (app_dir / "qt" / "telas" / "inteligencia_dialog.py").read_text(encoding="utf-8")


# ============================================================================
# POLIMENTO (pré-F12) — RG-53 na barra da Fábrica (teste-espelho da Mesa)
# ============================================================================

def test_barra_fabrica_720p_essenciais_e_estouro():
    """RG-53 portado: a janela da Fábrica cabe a 1280 (720p) — a barra não a
    prende na largura do conteúdo; apertando, os sacrificáveis colapsam no
    "···" (nada some) e os ESSENCIAIS (importar/modelo/exportar) ficam."""
    from PySide6.QtCore import Qt
    from PySide6.QtWidgets import QApplication

    from app.qt.telas.fabrica import FabricaTela
    QApplication.instance() or QApplication([])
    f = FabricaTela()
    f.setAttribute(Qt.WidgetAttribute.WA_DontShowOnScreen, True)
    f.show()
    QApplication.processEvents()

    # sem o conserto, o minimumSizeHint da barra empurraria o mínimo da
    # janela acima de 1280 (a lição da Mesa)
    assert f.minimumSizeHint().width() <= 1280

    def _reflow_em(largura):
        f.resize(largura, 720)
        f._barra_fabrica.resize(largura, f._barra_fabrica.height() or 44)
        f._reflow_barra()

    def _colapsados():
        return sum(1 for w, _r, _t in f._sacrificaveis if not w.isVisibleTo(f))

    # largo: a largura vem do CONTEÚDO REAL (sizeHints) — imune ao vazamento
    # de escala/fonte entre bancadas (a lição do teste da Mesa)
    from app.qt.design import tokens as tk
    esp = tk.ESP_2
    demanda = 2 * tk.ESP_3 + esp + f._mais_fabrica.sizeHint().width() + esp
    lay = f._barra_layout
    for i in range(lay.count()):
        w = lay.itemAt(i).widget()
        if w is not None and w is not f._mais_fabrica:
            demanda += w.sizeHint().width() + esp
    _reflow_em(demanda + 60)
    assert _colapsados() == 0
    assert not f._mais_fabrica.isVisibleTo(f)
    # estreito: sacrificáveis colapsam para o menu — e NADA some (cada
    # colapsado vira uma ação do "···")
    _reflow_em(700)
    n = _colapsados()
    assert n > 0
    assert f._mais_fabrica.isVisibleTo(f)
    assert len(f._mais_fabrica.menu().actions()) == n
    # essenciais nunca colapsam (não estão na lista de sacrifício)
    essenciais = {id(f.btn_exportar), id(f.combo_layout)}
    assert not (essenciais & {id(w) for w, _r, _t in f._sacrificaveis})
    f.close()


# ============================================================================
# CONSERTOS DA FROTA ADVERSARIAL (achados confirmados, corrigidos antes do selo)
# ============================================================================

def _renomear_no_banco(root, id_produto, novo_nome, nova_marca):
    from app.core.database import Database
    from app.core.models import Produto
    db = Database(root).init()
    try:
        with db.Session() as s:
            p = s.get(Produto, id_produto)
            p.nome_sanitizado, p.marca = novo_nome, nova_marca
            s.commit()
    finally:
        db.engine.dispose()


def test_excel_conflito_casa_por_chave_natural_nao_por_id(tmp_path):
    """[CRÍTICO — frota] o conflito casa por CHAVE NATURAL, não por id: se o
    produto for RENOMEADO entre analisar e aplicar (id igual, chave nova),
    USAR_PACOTE NÃO grava no produto errado — pula com aviso (I1/I2). Prova de
    mutação: o código antigo (s.get por id) corromperia o renomeado."""
    from openpyxl import load_workbook

    from app.core import excel_acervo as X
    from app.core.portabilidade import Decisao
    root = seeds.raiz(tmp_path, "loja")
    pid = seeds.add_produto(root, "Arroz 5kg", "Camil", "24.90", categoria="Mercearia")
    xlsx = X.exportar_acervo_xlsx(tmp_path / "a.xlsx", raiz=root)
    wb = load_workbook(str(xlsx))
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "Arroz 5kg":
            row[4].value = "19,90"
    wb.save(str(xlsx))
    a = X.analisar_planilha(xlsx, raiz=root)
    assert len(a.conflitos) == 1
    # o acervo MUDA entre analisar e aplicar: o id=pid vira outro produto
    _renomear_no_banco(root, pid, "Macarrão 500g", "Renata")
    dec = {c.id_decisao: Decisao.USAR_PACOTE for c in a.conflitos}
    rel = X.aplicar_importacao_planilha(a, dec, raiz=root)
    # o Macarrão (que herdou o id) NÃO foi corrompido com o preço do Arroz
    assert seeds.produto_por_chave(root, "Macarrão 500g", "Renata")["preco"] == "24.90"
    assert rel.avisos          # avisou que a identidade mudou (I2)


def test_excel_validade_data_real_roundtrip(tmp_path):
    """[ALTO — frota] célula de DATA real do Excel (datetime, não texto) é lida
    certo: o produto com validade não vira conflito fantasma. Prova de mutação:
    o _parse_data antigo devolvia None p/ 'YYYY-MM-DD 00:00:00' e a validade
    sumia calada."""
    from datetime import date as _date, datetime as _dt

    from openpyxl import load_workbook

    from app.core import excel_acervo as X
    from app.core.database import Database
    from app.core.models import Produto
    root = seeds.raiz(tmp_path, "loja")
    pid = seeds.add_produto(root, "Leite 1L", "Piracanjuba", "5.90",
                            categoria="Frios")
    db = Database(root).init()
    try:
        with db.Session() as s:
            s.get(Produto, pid).validade_item = _date(2026, 7, 6)
            s.commit()
    finally:
        db.engine.dispose()
    xlsx = X.exportar_acervo_xlsx(tmp_path / "a.xlsx", raiz=root)
    # o Excel guarda datas como DATETIME quando o operador digita/formata
    wb = load_workbook(str(xlsx))
    ws = wb.active
    idx_val = X.COLUNAS.index("Validade")
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "Leite 1L":
            row[idx_val].value = _dt(2026, 7, 6)     # célula de DATA real
    wb.save(str(xlsx))
    a = X.analisar_planilha(xlsx, raiz=root)
    assert "Leite 1L" in a.identicos          # a data casou — sem conflito fantasma
    assert a.conflitos == []


def test_impor_2em1_recusa_cartaz_grande():
    """[ALTO — frota] o 2-em-1 RECUSA (erro nominal) um cartaz maior que a
    metade da folha, em vez de cortar/sobrepor calado (I2)."""
    from app.rendering.imposicao import impor_2em1
    from app.rendering.units import mm_para_px
    dpi = 300
    a4 = Image.new("RGB", (round(mm_para_px(210, dpi)),
                          round(mm_para_px(297, dpi))), "white")   # A4 retrato
    with pytest.raises(ValueError, match="grande demais|A5|Dois por folha"):
        impor_2em1([a4, a4], dpi)


def test_fabrica_2em1_so_habilita_no_a5():
    """[ALTO — frota] o "Dois por folha" só fica habilitado para o A5/etiqueta —
    nunca deixa cortar um A4."""
    from PySide6.QtWidgets import QApplication

    from app.qt.telas.fabrica import FabricaTela
    QApplication.instance() or QApplication([])
    tela = FabricaTela()
    tela.combo_categoria  # noqa: B018 (garante construção)
    tela.combo_layout.setCurrentText("Meia folha A5 (148×210 mm)")
    assert tela.chk_2em1.isEnabled()
    tela.combo_layout.setCurrentText("A4 retrato (210×297 mm)")
    assert not tela.chk_2em1.isEnabled()
    assert not tela.chk_2em1.isChecked()    # desmarcou ao desabilitar
    tela.close()


def test_historico_de_preco_deduplica_por_edicao():
    """[MÉDIO — frota] o mesmo produto duplicado numa edição (mesmo produto_id,
    o "Duplicar item" só troca o uid) vira UM ponto na série, não dois na mesma
    data — coerente com o ranking."""
    from app.qt.telas import inteligencia as I
    ed = _ed("Quintou", datetime(2026, 7, 18, 10),
             [_it("Arroz", "24,90", pid=1), _it("Arroz (cópia)", "19,90", pid=1)])
    s = I.serie_de_um([ed], ("pid", 1))
    assert len(s["pontos"]) == 1             # 1 ponto por edição, não 2
    # e o ranking conta essa edição 1×
    assert I.ranking_ofertados([ed])[0]["edicoes"] == 1
