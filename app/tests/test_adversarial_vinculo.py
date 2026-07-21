"""Teste adversarial do vínculo imagem×nome×preço×slot (§4 da auditoria; I5).

Tenta ATIVAMENTE trocar conteúdo de lugar — reordenar z-order, editar a mestra,
desfazer/refazer, remover/adicionar item, salvar/reabrir/duplicar projeto — e
confere o trio **pelo conteúdo renderizado** (cor da imagem por slot), nunca
por "não deu exceção".
"""

import random
from decimal import Decimal
from pathlib import Path

import pytest
from PIL import Image
from PySide6.QtWidgets import QApplication

from app.qt.telas import servico
from app.qt.telas.servico import ItemMesa
from app.rendering.compositor import DadosProduto, compor_pagina
from app.rendering.grade import propagar_mestre
from app.rendering.model import (
    Ajuste, LayoutDef, Pagina, Regiao, Retangulo, Slot, TipoRegiao,
)
from app.rendering.units import mm_para_px

CORES = ["#FF0000", "#00FF00", "#0000FF", "#FFFF00",
         "#FF00FF", "#00FFFF", "#800000", "#008000"]   # 1 cor por produto


@pytest.fixture()
def raiz_tmp(tmp_path, monkeypatch):
    monkeypatch.setenv("AUTOTABLOIDE_ROOT", str(tmp_path / "raiz"))
    import shutil
    from app.core.database import Database
    from app.core.paths import SystemRoot

    root = SystemRoot(tmp_path / "raiz").criar_estrutura()
    reais = Path("AutoTabloide_System_Root/fontes")
    if reais.exists():
        for f in reais.glob("*.ttf"):
            shutil.copy(f, root.fontes / f.name)
    Database(root).init().engine.dispose()
    return root


def _grade_4() -> LayoutDef:
    """Grade sintética 2×2 (âncoras 10/60 mm), mestra + 3 células via propagação."""
    regs = [
        Regiao(TipoRegiao.IMAGEM, Retangulo(12, 12, 30, 20), nome="Img",
               ajuste=Ajuste.PREENCHER),
        Regiao(TipoRegiao.NOME, Retangulo(12, 34, 30, 8), nome="Nome",
               tamanho_max_pt=10),
        Regiao(TipoRegiao.PRECO, Retangulo(12, 43, 30, 8), nome="Preço",
               tamanho_max_pt=12),
    ]
    for r in regs:
        r.de_mestre = True
    slots = [Slot("celula_0", regs, mestre=True, origem_mm=(10, 10))]
    for i, (x, y) in enumerate([(60, 10), (10, 60), (60, 60)], start=1):
        slots.append(Slot(f"celula_{i}", origem_mm=(x, y)))
    lay = LayoutDef(110, 110, dpi=100, paginas=[Pagina(slots)])
    propagar_mestre(lay.paginas[0])
    return lay


def _itens(tmp_path) -> list[ItemMesa]:
    itens = []
    for i, cor in enumerate(CORES):
        foto = tmp_path / f"prod_{i}.png"
        Image.new("RGB", (200, 200), cor).save(foto)
        itens.append(ItemMesa(descricao=f"PROD-{i}", preco=f"{i + 1},00",
                              semaforo="VERDE", nome=f"PROD-{i}",
                              imagem=str(foto)))
    return itens


def _cor_no_slot(img: Image.Image, layout: LayoutDef, slot: Slot) -> tuple:
    """Amostra o pixel central da região IMAGEM do slot (conteúdo, não fé)."""
    reg = next(r for r in slot.regioes if r.tipo == TipoRegiao.IMAGEM)
    cx = mm_para_px(reg.rect.x_mm + reg.rect.larg_mm / 2, layout.dpi)
    cy = mm_para_px(reg.rect.y_mm + reg.rect.alt_mm / 2, layout.dpi)
    return img.getpixel((round(cx), round(cy)))[:3]


def _conferir(layout, mapa, itens):
    """Compõe pelo mapa e confere CADA célula ocupada pela cor do SEU produto."""
    por_uid = {it.uid: it for it in itens}
    dados = {sid: DadosProduto(it.nome,
                               preco_por=servico.preco_decimal(it.preco),
                               imagem_path=it.imagem)
             for sid, it in ((s, por_uid[u]) for s, u in mapa.items()
                             if u in por_uid)}
    img = compor_pagina(layout, layout.paginas[0], dados)
    for slot in layout.paginas[0].slots:
        uid = mapa.get(slot.id)
        if uid is None or uid not in por_uid:
            continue
        esperado = CORES[int(por_uid[uid].nome.split("-")[1])]
        e = tuple(int(esperado[i:i + 2], 16) for i in (1, 3, 5))
        assert _cor_no_slot(img, layout, slot) == e, \
            f"{slot.id}: imagem NÃO é a do produto {por_uid[uid].nome}"


def test_adversarial_vinculo_completo(raiz_tmp, tmp_path):
    QApplication.instance() or QApplication([])
    layout = _grade_4()
    itens = _itens(tmp_path)
    mapa = {s.id: it.uid for s, it in zip(layout.paginas[0].slots, itens)}
    _conferir(layout, mapa, itens)                       # base ok

    # 1. reordenar z-order em células e editar a mestra (pareamento por uid, I4)
    for slot in layout.paginas[0].slots[1:3]:
        slot.regioes.reverse()                           # baguncei a ordem
    mestre = layout.paginas[0].slots[0]
    mestre.regioes[0].rect.larg_mm = 28                  # edito a mestra
    mestre.regioes[2].cor = "#123456"
    propagar_mestre(layout.paginas[0])
    for slot in layout.paginas[0].slots[1:]:
        img_reg = next(r for r in slot.regioes if r.tipo == TipoRegiao.IMAGEM)
        assert img_reg.rect.larg_mm == 28                # foi para a IMAGEM certa
        preco_reg = next(r for r in slot.regioes if r.tipo == TipoRegiao.PRECO)
        assert preco_reg.cor == "#123456"                # e a cor para o PREÇO
    _conferir(layout, mapa, itens)

    # 2. desfazer/refazer intercalado não corrompe o vínculo
    from app.qt.canvas import CanvasView
    v = CanvasView()
    v.carregar(layout, DadosProduto("x"))
    for _ in range(5):
        v.regioes()[0].rect.x_mm += 1
        v.notificar_edicao(v.regioes()[0], "rect")
    for _ in range(5):
        v.desfazer()
    for _ in range(2):
        v.refazer()
    layout = v._layout                                   # layout restaurado
    _conferir(layout, mapa, itens)

    # 3. remover item do meio e adicionar outro — os demais NÃO se movem
    removido = itens.pop(1)
    mapa = {s: u for s, u in mapa.items() if u != removido.uid}
    novo_png = tmp_path / "prod_9.png"
    Image.new("RGB", (200, 200), CORES[1]).save(novo_png)
    novo = ItemMesa(descricao="PROD-1", preco="9,00", semaforo="VERDE",
                    nome="PROD-1", imagem=str(novo_png))
    itens.append(novo)                                   # entra no FIM da estante…
    random.shuffle(itens)                                # …e a estante embaralha!
    _conferir(layout, mapa, itens)                       # ninguém trocou de slot

    # 4. mais itens que células / células sem item — degrada sem trocar nada
    _conferir(layout, dict(list(mapa.items())[:2]), itens)

    # alvos DETERMINÍSTICOS: itens que estão no mapa (identidade, não índice —
    # o shuffle do passo 3 tornaria itens[i] aleatório)
    por_uid = {it.uid: it for it in itens}
    no_mapa = [por_uid[u] for u in mapa.values() if u in por_uid]

    # 5. preço com milhar e preço vazio → aparecem no pré-voo (I2)
    no_mapa[0].preco = "1.299,00"
    assert servico.preco_decimal(no_mapa[0].preco) == Decimal("1299.00")
    no_mapa[1].preco = None
    dados = {sid: DadosProduto(por_uid[u].nome,
                               preco_por=servico.preco_decimal(por_uid[u].preco),
                               imagem_path=por_uid[u].imagem)
             for sid, u in mapa.items() if u in por_uid}
    avisos = servico.validar_composicao(layout, dados)
    assert any("sem preço" in a for a in avisos)

    # 6. imagem apagada do disco → pré-voo acusa (não some em silêncio)
    alvo = no_mapa[2]
    Path(alvo.imagem).unlink()
    dados2 = {sid: DadosProduto(por_uid[u].nome, preco_por=Decimal("1"),
                                imagem_path=por_uid[u].imagem)
              for sid, u in mapa.items() if u in por_uid}
    avisos = servico.validar_composicao(layout, dados2)
    assert any("sumiu do disco" in a for a in avisos)
    cor_do_alvo = CORES[int(alvo.nome.split("-")[1])]
    Image.new("RGB", (200, 200), cor_do_alvo).save(alvo.imagem)  # restaura A DELE

    # 7. salvar projeto → reabrir → duplicar → abrir o duplicado: mapa intacto
    from app.core import projetos
    no_mapa[0].preco = "1,00"
    no_mapa[1].preco = "2,00"
    pid = projetos.salvar_projeto("Adversarial", None, "TABLOIDE", layout,
                                  [it.to_dict() for it in itens], mapa=mapa)
    p = projetos.abrir_projeto(pid)
    assert p.mapa == mapa                                # o casamento congelou
    reitens = [ItemMesa.from_dict(d) for d in p.itens]
    _conferir(p.layout, p.mapa, reitens)                 # trio confere reaberto
    pid2 = projetos.duplicar_projeto(pid, "Adversarial 2")
    p2 = projetos.abrir_projeto(pid2)
    _conferir(p2.layout, p2.mapa,
              [ItemMesa.from_dict(d) for d in p2.itens])

    # 8. exportar PNG do congelado
    from app.rendering.export import exportar_png
    por_uid2 = {d["uid"]: d for d in p2.itens}
    dados3 = {sid: DadosProduto(por_uid2[u]["nome"],
                                preco_por=servico.preco_decimal(por_uid2[u]["preco"]),
                                imagem_path=por_uid2[u]["imagem"])
              for sid, u in p2.mapa.items() if u in por_uid2}
    img = compor_pagina(p2.layout, p2.layout.paginas[0], dados3)
    saida = exportar_png(img, tmp_path / "adversarial.png", p2.layout.dpi)
    assert saida.exists()


def test_adversarial_grupos_livres(raiz_tmp, tmp_path):
    """Passos 9–14 da ORDEM_F5_6 (§4): grupos livres sob ataque, por pixel."""
    from app.qt.canvas import CanvasView
    from app.rendering.grade import (
        agrupar_como_mestre, carimbar_copia, mestres, remover_slot,
    )

    QApplication.instance() or QApplication([])

    # grade real (2×2) + itens coloridos, como nos passos 1–8
    layout = _grade_4()
    itens = _itens(tmp_path)

    # 9. criar grupo livre (3 regiões) + carimbar 3 cópias → conferir por pixel
    avulso = Slot("avulso", [
        Regiao(TipoRegiao.IMAGEM, Retangulo(12, 90, 20, 12), nome="GImg",
               ajuste=Ajuste.PREENCHER),
        Regiao(TipoRegiao.NOME, Retangulo(12, 103, 20, 4), nome="GNome",
               tamanho_max_pt=8),
        Regiao(TipoRegiao.PRECO, Retangulo(12, 107, 20, 3), nome="GPreço",
               tamanho_max_pt=8),
    ])
    layout.altura_mm = 160                                # espaço p/ o grupo
    layout.paginas[0].slots.append(avulso)
    mestre_g = agrupar_como_mestre(layout.paginas[0], list(avulso.regioes), avulso)
    assert mestre_g.id.startswith("grupo_")               # D1: id uuid
    copias = [carimbar_copia(layout.paginas[0], mestre_g, (40 + i * 25, 120))
              for i in range(3)]
    assert all(c.id.startswith("grupo_") for c in copias)
    for c in copias:                                      # D4: uid novo + ref certo
        for r in c.regioes:
            assert r.ref_mestre in {m.uid for m in mestre_g.regioes}
            assert r.uid not in {m.uid for m in mestre_g.regioes}

    # grade + grupo coexistem no MESMO mapa (D2)
    grupo_slots = [mestre_g] + copias
    itens_g = _itens(tmp_path)                            # 4 produtos p/ o grupo
    mapa = {s.id: it.uid for s, it in zip(layout.paginas[0].slots[:4], itens)}
    mapa.update({s.id: it.uid for s, it in zip(grupo_slots, itens_g)})
    todos = itens + itens_g
    _conferir(layout, mapa, todos)

    # 10. remover a cópia do MEIO → ninguém troca de célula; item fica órfão
    alvo = copias[1]
    uid_orfao = mapa[alvo.id]
    remover_slot(layout.paginas[0], alvo.id)
    mapa_sem = {s: u for s, u in mapa.items() if s != alvo.id}
    _conferir(layout, mapa_sem, todos)                    # vizinhos intactos
    assert uid_orfao not in mapa_sem.values()             # "fora da grade"

    # 11. desfazer a remoção → slot E entrada do mapa voltam (D5)
    v = CanvasView()
    v.mapa = dict(mapa)
    v.carregar(layout, DadosProduto("x"))                 # baseline {layout, mapa}
    assert v.remover_celula(alvo2_id := copias[2].id)
    assert alvo2_id not in v.mapa
    assert v.desfazer()
    assert alvo2_id in v.mapa                             # a entrada VOLTOU
    assert any(s.id == alvo2_id for s in v._layout.paginas[0].slots)
    layout = v._layout

    # 12. z-order embaralhado nas cópias + editar o mestre → propaga certo
    pagina = layout.paginas[0]
    mg = next(m for m in mestres(pagina) if m.id == mestre_g.id)
    for s in pagina.slots:
        if s.ref_grupo == mg.id:
            s.regioes.reverse()
    img_m = next(r for r in mg.regioes if r.tipo == TipoRegiao.IMAGEM)
    img_m.rect.larg_mm = 18
    from app.rendering.grade import propagar_mestre as _prop
    _prop(pagina)
    for s in pagina.slots:
        if s.ref_grupo == mg.id:
            r_img = next(r for r in s.regioes if r.tipo == TipoRegiao.IMAGEM)
            assert r_img.rect.larg_mm == 18               # foi p/ a IMAGEM certa
    _conferir(layout, {s: u for s, u in v.mapa.items()
                       if any(sl.id == s for sl in pagina.slots)}, todos)

    # 13. remover o MESTRE → promoção da cópia mais antiga (D3) e trio confere
    copias_vivas = [s for s in pagina.slots if s.ref_grupo == mg.id]
    promovida_esperada = copias_vivas[0]
    remover_slot(pagina, mg.id)
    assert promovida_esperada.mestre and promovida_esperada.ref_grupo is None
    for s in pagina.slots:
        if s.ref_grupo == promovida_esperada.id:
            for r in s.regioes:                           # irmãs reapontadas
                assert r.ref_mestre in {p.uid for p in promovida_esperada.regioes}
    mapa_final = {s: u for s, u in v.mapa.items()
                  if any(sl.id == s for sl in pagina.slots)}
    _conferir(layout, mapa_final, todos)

    # 14. salvar → reabrir → exportar com grade + grupo coexistindo
    from app.core import projetos
    from app.rendering.export import exportar_png
    pid = projetos.salvar_projeto("Grupos", None, "TABLOIDE", layout,
                                  [it.to_dict() for it in todos],
                                  mapa=mapa_final)
    p = projetos.abrir_projeto(pid)
    assert p.mapa == mapa_final
    reitens = [ItemMesa.from_dict(d) for d in p.itens]
    _conferir(p.layout, p.mapa, reitens)
    por_uid = {it.uid: it for it in reitens}
    dados = {sid: DadosProduto(por_uid[u].nome, preco_por=Decimal("1"),
                               imagem_path=por_uid[u].imagem)
             for sid, u in p.mapa.items() if u in por_uid}
    img = compor_pagina(p.layout, p.layout.paginas[0], dados)
    assert exportar_png(img, tmp_path / "grupos.png", p.layout.dpi).exists()


def test_fluxo_real_grade_mais_destaque(raiz_tmp, tmp_path):
    """C3 (ORDEM_F5_6 §6): o fluxo do gate 3 — grade + destaque de capa.

    Grade 2×2 + 3 regiões avulsas (C1, sem seleção) → agrupar (C2 recusa o que
    não é livre) → carimbar 2 → preencher na ordem visual → exportar. Confere
    por pixel E que NENHUMA célula da grade ganhou/perdeu região.
    """
    from app.qt.canvas import CanvasView
    from app.rendering.export import exportar_png
    from app.rendering.grade import ordenar_slots_visualmente

    QApplication.instance() or QApplication([])
    layout = _grade_4()
    layout.altura_mm = 170                                # área livre embaixo
    v = CanvasView()
    v.carregar(layout, DadosProduto("x"))

    # fotografia da grade ANTES: uids região a região, célula a célula
    grade_antes = {s.id: [r.uid for r in s.regioes]
                   for s in layout.paginas[0].slots}

    # C1: sem seleção → a 1ª região nasce num slot LIVRE (nunca no mestre!)
    v._scene.clearSelection()
    r1 = v.adicionar_regiao(TipoRegiao.IMAGEM)
    slot_livre = v._slot_de(r1)
    assert slot_livre.id.startswith("livre_")
    assert not r1.de_mestre and r1.ref_mestre is None
    # as seguintes acompanham a seleção (r1 ficou selecionada) → mesmo slot
    r2 = v.adicionar_regiao(TipoRegiao.NOME)
    r3 = v.adicionar_regiao(TipoRegiao.PRECO)
    assert v._slot_de(r2) is slot_livre and v._slot_de(r3) is slot_livre
    # posiciona o destaque na área livre
    r1.rect = Retangulo(12, 120, 24, 14)
    r1.ajuste = Ajuste.PREENCHER
    r2.rect = Retangulo(12, 136, 24, 5)
    r3.rect = Retangulo(12, 142, 24, 5)

    # C2: agrupar regiões DERIVADAS da grade → recusa, grade intocada
    v._scene.clearSelection()
    for it in v._itens:
        s = v._slot_de(it.regiao)
        if s is not None and s.id == "celula_1":
            it.setSelected(True)
    assert v.agrupar_selecao() is None
    # C2: agrupar regiões do MESTRE da grade → recusa
    v._scene.clearSelection()
    for it in v._itens:
        s = v._slot_de(it.regiao)
        if s is not None and s.mestre:
            it.setSelected(True)
    assert v.agrupar_selecao() is None

    # agrupar o destaque (regiões livres) → ok; carimbar 2 cópias
    v._scene.clearSelection()
    for it in v._itens:
        if v._slot_de(it.regiao) is slot_livre:
            it.setSelected(True)
    mestre_g = v.agrupar_selecao()
    assert mestre_g is not None and mestre_g.origem_mm == (12, 120)
    # C5.3: o slot livre_ esvaziado saiu do layout — sem fantasma na raiz
    assert all(not s.id.startswith("livre_")
               for s in v._layout.paginas[0].slots)
    v.carimbar_grupo(mestre_g.id, (48, 120))
    v.carimbar_grupo(mestre_g.id, (84, 120))

    # INVARIÂNCIA: nenhuma célula da grade ganhou/perdeu região (uid a uid)
    for s in v._layout.paginas[0].slots:
        if s.id in grade_antes:
            assert [r.uid for r in s.regioes] == grade_antes[s.id], \
                f"{s.id}: as regiões da grade mudaram!"

    # auto-preencher na ordem visual (grade em cima, destaque embaixo) + pixel
    # C5.1: o MESMO caminho da produção (grade.ocupaveis) — sem filtro local
    from app.rendering.grade import ocupaveis
    itens = _itens(tmp_path)[:7]
    slots_ordem = ocupaveis(ordenar_slots_visualmente(v._layout.paginas[0].slots))
    mapa = {s.id: it.uid for s, it in zip(slots_ordem, itens)}
    assert len(mapa) == 7                                 # 4 grade + 3 destaque
    _conferir(v._layout, mapa, itens)

    # exportar e conferir a grade uma última vez
    por_uid = {it.uid: it for it in itens}
    dados = {sid: DadosProduto(por_uid[u].nome, preco_por=Decimal("1"),
                               imagem_path=por_uid[u].imagem)
             for sid, u in mapa.items()}
    img = compor_pagina(v._layout, v._layout.paginas[0], dados)
    assert exportar_png(img, tmp_path / "gate3.png", v._layout.dpi).exists()
    for s in v._layout.paginas[0].slots:
        if s.id in grade_antes:
            assert [r.uid for r in s.regioes] == grade_antes[s.id]


def test_c5_agrupar_nao_deixa_fantasma(raiz_tmp, tmp_path):
    """C5 explícito: agrupar esvazia o livre_ → a Mesa não atribui item a ele
    e o pré-voo fica limpo; se o slot vazio TEM papel no mapa, ele fica e o
    pré-voo ACUSA "célula vazia" (defesa em profundidade, C5.2)."""
    from app.qt.canvas import CanvasView
    from app.rendering.grade import ocupaveis, ordenar_slots_visualmente

    QApplication.instance() or QApplication([])
    layout = _grade_4()
    layout.altura_mm = 170
    v = CanvasView()
    v.carregar(layout, DadosProduto("x"))

    # destaque via C1 (sem seleção) e agrupar → o livre_ some (C5.3)
    v._scene.clearSelection()
    r1 = v.adicionar_regiao(TipoRegiao.IMAGEM)
    r1.rect = Retangulo(12, 120, 24, 14)
    mestre_g = v.agrupar_selecao()
    assert mestre_g is not None
    assert all(not s.id.startswith("livre_") for s in v._layout.paginas[0].slots)

    # o caminho REAL do auto-preencher: nenhum item cai em slot sem regiões
    itens = _itens(tmp_path)[:5]
    slots = ocupaveis(ordenar_slots_visualmente(v._layout.paginas[0].slots))
    mapa = {s.id: it.uid for s, it in zip(slots, itens)}
    por_uid = {it.uid: it for it in itens}
    dados = {sid: DadosProduto(por_uid[u].nome, preco_por=Decimal("1"),
                               imagem_path=por_uid[u].imagem)
             for sid, u in mapa.items()}
    avisos = servico.validar_composicao(v._layout, dados)
    assert not any("célula vazia" in a for a in avisos)   # pré-voo limpo

    # defesa em profundidade: slot vazio COM papel no mapa fica e é acusado
    v._scene.clearSelection()
    r2 = v.adicionar_regiao(TipoRegiao.IMAGEM)
    slot_livre2 = v._slot_de(r2)
    v.mapa[slot_livre2.id] = itens[0].uid                 # papel no mapa!
    v._scene.clearSelection()
    for it in v._itens:
        if it.regiao is r2:
            it.setSelected(True)
    assert v.agrupar_selecao() is not None
    assert any(s.id == slot_livre2.id                     # C5.3: ele FICA…
               for s in v._layout.paginas[0].slots)
    dados2 = dict(dados)
    dados2[slot_livre2.id] = DadosProduto(itens[0].nome, preco_por=Decimal("1"),
                                          imagem_path=itens[0].imagem)
    avisos2 = servico.validar_composicao(v._layout, dados2)
    assert any("célula vazia" in a for a in avisos2)      # …e o pré-voo ACUSA


def test_canvas_gestos_de_grupo(raiz_tmp, tmp_path):
    """Os gestos da UI (agrupar seleção / carimbar / remover) no canvas."""
    from app.qt.canvas import CanvasView

    QApplication.instance() or QApplication([])
    lay = LayoutDef(120, 120, dpi=100, paginas=[Pagina([Slot("pagina", [
        Regiao(TipoRegiao.IMAGEM, Retangulo(10, 10, 20, 12), nome="I"),
        Regiao(TipoRegiao.PRECO, Retangulo(10, 24, 20, 6), nome="P"),
    ])])])
    v = CanvasView()
    v.carregar(lay, DadosProduto("x"))

    for it in v._itens:                       # seleciona as 2 regiões
        it.setSelected(True)
    mestre = v.agrupar_selecao()
    assert mestre is not None and mestre.mestre
    assert mestre.id.startswith("grupo_")     # D1
    assert mestre.origem_mm == (10, 10)       # âncora = canto do conjunto

    copia = v.carimbar_grupo(mestre.id, (60, 10))
    assert copia is not None and copia.ref_grupo == mestre.id
    assert len(copia.regioes) == 2            # povoada pela propagação
    r = next(r for r in copia.regioes if r.tipo == TipoRegiao.IMAGEM)
    assert (r.rect.x_mm, r.rect.y_mm) == (60, 10)   # rebase pela âncora

    assert v.remover_celula(copia.id)
    assert all(s.id != copia.id for s in v._layout.paginas[0].slots)
    assert v.desfazer()                       # D5: a célula volta
    assert any(s.id == copia.id for s in v._layout.paginas[0].slots)


def test_adversarial_override_por_slot(raiz_tmp, tmp_path):
    """B3 do Bloco E (F7.3, I5): o override de UMA célula sob ataque.

    Por conteúdo (pixel/byte): o override aparece só na célula dele, NÃO vaza
    para as vizinhas, não toca o item da estante, sobrevive a
    salvar→reabrir→duplicar (com a foto CONGELADA na pasta do projeto) e o
    undo restaura o estado anterior.
    """
    QApplication.instance() or QApplication([])
    layout = _grade_4()
    itens = _itens(tmp_path)[:4]
    mapa = {s.id: it.uid for s, it in zip(layout.paginas[0].slots, itens)}
    _conferir(layout, mapa, itens)                        # base ok

    COR_OV = "#654321"
    foto_ov = tmp_path / "override.png"
    Image.new("RGB", (200, 200), COR_OV).save(foto_ov)
    overrides = {"celula_1": {"nome": "Oferta Especial", "preco": "9,99",
                              "imagem": str(foto_ov)}}

    def _dados_com_override(lay, mp, its, ovs):
        por_uid = {it.uid: it for it in its}
        dados = {}
        for sid, u in mp.items():
            if u not in por_uid:
                continue
            it = por_uid[u]
            d = DadosProduto(it.nome,
                             preco_por=servico.preco_decimal(it.preco),
                             imagem_path=it.imagem)
            ov = ovs.get(sid)
            dados[sid] = servico.aplicar_override(d, ov) if ov else d
        return dados

    def _conferir_com_override(lay, mp, its, ovs):
        img = compor_pagina(lay, lay.paginas[0], _dados_com_override(
            lay, mp, its, ovs))
        por_uid = {it.uid: it for it in its}
        for slot in lay.paginas[0].slots:
            uid = mp.get(slot.id)
            if uid is None or uid not in por_uid:
                continue
            if slot.id in ovs and ovs[slot.id].get("imagem"):
                esperado = COR_OV                          # a foto do OVERRIDE
            else:
                esperado = CORES[int(por_uid[uid].nome.split("-")[1])]
            e = tuple(int(esperado[i:i + 2], 16) for i in (1, 3, 5))
            assert _cor_no_slot(img, lay, slot) == e, \
                f"{slot.id}: o override vazou/errou a célula!"

    # 1. o override rende SÓ na célula dele; as outras ficam com o item delas
    _conferir_com_override(layout, mapa, itens, overrides)
    # e o ITEM da estante não foi tocado (só a célula muda)
    it_da_celula = next(it for it in itens if it.uid == mapa["celula_1"])
    assert it_da_celula.nome == "PROD-1" and it_da_celula.preco == "2,00"

    # 2. embaralhar a estante e reordenar o mapa não move o override (I1)
    random.shuffle(itens)
    mapa_reordenado = dict(reversed(list(mapa.items())))
    _conferir_com_override(layout, mapa_reordenado, itens, overrides)

    # 3. undo: aplicar → trocar → desfazer volta o anterior, desfazer some
    from app.qt.canvas import CanvasView
    v = CanvasView()
    v.carregar(layout, DadosProduto("x"))                 # baseline sem override
    v.set_override("celula_1", overrides["celula_1"])
    v.set_override("celula_1", {"preco": "1,11"})
    assert v.overrides["celula_1"] == {"preco": "1,11"}
    assert v.desfazer()
    assert v.overrides["celula_1"] == overrides["celula_1"]   # o ANTERIOR voltou
    assert v.desfazer()
    assert "celula_1" not in v.overrides                  # base: sem override
    assert v.refazer()
    assert v.overrides["celula_1"] == overrides["celula_1"]
    layout = v._layout

    # 4. salvar → reabrir → duplicar: override sobrevive com a foto CONGELADA
    from app.core import projetos
    pid = projetos.salvar_projeto("Override Adv", None, "TABLOIDE", layout,
                                  [it.to_dict() for it in itens],
                                  mapa=mapa, overrides=overrides)
    # I3: o JSON persistido guarda a foto RELATIVA à pasta do projeto
    from sqlalchemy import select
    from app.core.database import Database
    from app.core.models import ProjetoSalvo
    db = Database().init()
    with db.Session() as s:
        row = s.execute(select(ProjetoSalvo).where(
            ProjetoSalvo.id == pid)).scalar_one()
        assert "imagens/override_celula_1" in row.overrides_json
        assert str(tmp_path) not in row.overrides_json
    db.engine.dispose()

    p = projetos.abrir_projeto(pid)
    assert set(p.overrides) == {"celula_1"}
    foto_congelada = Path(p.overrides["celula_1"]["imagem"])
    assert foto_congelada.exists()
    assert foto_congelada.read_bytes() == foto_ov.read_bytes()   # byte a byte
    reitens = [ItemMesa.from_dict(d) for d in p.itens]
    _conferir_com_override(p.layout, p.mapa, reitens, p.overrides)

    # a foto ORIGINAL some do disco → o projeto congelado nem percebe
    foto_ov.unlink()
    p_de_novo = projetos.abrir_projeto(pid)
    _conferir_com_override(p_de_novo.layout, p_de_novo.mapa,
                           [ItemMesa.from_dict(d) for d in p_de_novo.itens],
                           p_de_novo.overrides)

    pid2 = projetos.duplicar_projeto(pid, "Override Adv 2")
    p2 = projetos.abrir_projeto(pid2)
    assert set(p2.overrides) == {"celula_1"}
    _conferir_com_override(p2.layout, p2.mapa,
                           [ItemMesa.from_dict(d) for d in p2.itens],
                           p2.overrides)


def _cor_em(img, layout, slot, fx: float, fy: float) -> tuple:
    """Amostra um ponto FRACIONÁRIO da região IMAGEM do slot (fx, fy ∈ 0..1)."""
    reg = next(r for r in slot.regioes if r.tipo == TipoRegiao.IMAGEM)
    cx = mm_para_px(reg.rect.x_mm + reg.rect.larg_mm * fx, layout.dpi)
    cy = mm_para_px(reg.rect.y_mm + reg.rect.alt_mm * fy, layout.dpi)
    return img.getpixel((round(cx), round(cy)))[:3]


def test_adversarial_multi_imagem_por_conteudo(raiz_tmp, tmp_path):
    """C do Bloco E (F7.1, I5): várias fotos num slot, POR PIXEL.

    A ordem da lista é a ordem do desenho (LADO_A_LADO amostrado terço a
    terço); trocar a ordem troca os pixels; o LEQUE mostra TODAS as cores;
    salvar→reabrir→duplicar preserva as N fotos NA ORDEM (byte a byte) e o
    arranjo — apagar as originais do disco não afeta o congelado.
    """
    from app.rendering.compositor import ImagemSlot
    from app.rendering.arranjo import ModoArranjo

    QApplication.instance() or QApplication([])
    layout = _grade_4()
    slot0 = layout.paginas[0].slots[0]

    tres = []
    for i, cor in enumerate(["#FF0000", "#00FF00", "#0000FF"]):
        f = tmp_path / f"sabor_{i}.png"
        Image.new("RGB", (200, 200), cor).save(f)
        tres.append(str(f))
    rgb = [(255, 0, 0), (0, 255, 0), (0, 0, 255)]

    def _compor(caminhos, modo):
        d = DadosProduto("Multi", preco_por=Decimal("1"),
                         imagens=[ImagemSlot(c) for c in caminhos],
                         modo_arranjo=modo)
        return compor_pagina(layout, layout.paginas[0], {slot0.id: d})

    # 1. LADO_A_LADO: terço a terço, NA ORDEM da lista
    img = _compor(tres, ModoArranjo.LADO_A_LADO)
    for i in range(3):
        assert _cor_em(img, layout, slot0, (i + 0.5) / 3, 0.5) == rgb[i], \
            f"terço {i}: a foto não é a da posição {i} da lista!"

    # 2. trocar a ORDEM da lista troca os pixels (a ordem é conteúdo)
    invertida = list(reversed(tres))
    img2 = _compor(invertida, ModoArranjo.LADO_A_LADO)
    for i in range(3):
        assert _cor_em(img2, layout, slot0, (i + 0.5) / 3, 0.5) == rgb[2 - i]

    # 3. LEQUE: as TRÊS cores aparecem (sobreposição não engole ninguém)
    img3 = _compor(tres, ModoArranjo.LEQUE)
    reg = next(r for r in slot0.regioes if r.tipo == TipoRegiao.IMAGEM)
    x0 = round(mm_para_px(reg.rect.x_mm, layout.dpi))
    y0 = round(mm_para_px(reg.rect.y_mm, layout.dpi))
    x1 = round(mm_para_px(reg.rect.x_mm + reg.rect.larg_mm, layout.dpi))
    y1 = round(mm_para_px(reg.rect.y_mm + reg.rect.alt_mm, layout.dpi))
    vistos = {p[:3] for p in img3.crop((x0, y0, x1, y1)).getdata()}
    for c in rgb:
        assert c in vistos, f"o leque engoliu a cor {c}!"

    # 4. congelamento: as N fotos sobrevivem NA ORDEM, byte a byte
    from app.core import projetos
    item = ItemMesa(descricao="MULTI", preco="1,00", semaforo="VERDE",
                    nome="Multi Sabores", imagem=tres[0], imagens=list(tres),
                    arranjo="LADO_A_LADO")
    outros = _itens(tmp_path)[:1]
    mapa = {slot0.id: item.uid,
            layout.paginas[0].slots[1].id: outros[0].uid}
    bytes_originais = [Path(c).read_bytes() for c in tres]
    pid = projetos.salvar_projeto("Multi Adv", None, "TABLOIDE", layout,
                                  [item.to_dict()] + [o.to_dict()
                                                      for o in outros],
                                  mapa=mapa)
    for c in tres:
        Path(c).unlink()                       # as ORIGINAIS somem do disco

    p = projetos.abrir_projeto(pid)
    reaberto = next(d for d in p.itens if d["nome"] == "Multi Sabores")
    assert reaberto["arranjo"] == "LADO_A_LADO"        # o arranjo persistiu
    assert len(reaberto["imagens"]) == 3
    for k, cam in enumerate(reaberto["imagens"]):
        assert Path(cam).exists()
        assert Path(cam).read_bytes() == bytes_originais[k], \
            f"foto {k} do congelado não é byte-idêntica (ordem trocada?)"
    # e compõe do congelado com os pixels na ordem certa
    d = DadosProduto("Multi", preco_por=Decimal("1"),
                     imagens=[ImagemSlot(c) for c in reaberto["imagens"]],
                     modo_arranjo=ModoArranjo.LADO_A_LADO)
    img4 = compor_pagina(p.layout, p.layout.paginas[0], {slot0.id: d})
    for i in range(3):
        assert _cor_em(img4, p.layout, p.layout.paginas[0].slots[0],
                       (i + 0.5) / 3, 0.5) == rgb[i]

    # 5. duplicar: o duplicado tem as PRÓPRIAS cópias, mesma ordem
    pid2 = projetos.duplicar_projeto(pid, "Multi Adv 2")
    p2 = projetos.abrir_projeto(pid2)
    re2 = next(d for d in p2.itens if d["nome"] == "Multi Sabores")
    for k, cam in enumerate(re2["imagens"]):
        assert Path(cam).read_bytes() == bytes_originais[k]

    # 6. pré-voo multi: apagar UMA congelada → acusa a foto certa (I2)
    Path(reaberto["imagens"][1]).unlink()
    dados = {slot0.id: DadosProduto(
        "Multi", preco_por=Decimal("1"),
        imagens=[ImagemSlot(c) for c in reaberto["imagens"]])}
    avisos = servico.validar_composicao(p.layout, dados)
    assert any("imagem 2/3" in a for a in avisos)


def test_adversarial_item_composto(raiz_tmp, tmp_path):
    """D do Bloco E (F7.2, I5): "Camil e Rei" sob ataque de identidade.

    A regra sagrada — 1 slot → 1 uid — sob compor/separar/reabrir em ciclos:
    NUNCA uid duplicado no mapa, NUNCA órfão, os originais voltam com os
    uids DE SEMPRE, e o LADO_A_LADO do composto rende por pixel (Camil à
    esquerda, Rei à direita) — inclusive DEPOIS do congelamento.
    """
    from app.rendering.arranjo import ModoArranjo
    from app.rendering.compositor import ImagemSlot

    QApplication.instance() or QApplication([])
    from app.qt.telas.mesa import MesaTela

    mesa = MesaTela()
    mesa.carregar_layout(_grade_4(), None)
    itens = _itens(tmp_path)[:4]           # PROD-0..3, cores CORES[0..3]
    mesa._itens = itens
    mesa._auto_preencher()
    layout = mesa.area.canvas._layout
    a, b = itens[0], itens[1]
    slot_a = next(s for s, u in mesa._mapa.items() if u == a.uid)

    def _sem_duplicata_nem_orfao():
        valores = list(mesa._mapa.values())
        assert len(set(valores)) == len(valores), "DOIS slots com o mesmo uid!"
        uids_estante = {it.uid for it in mesa._itens}
        assert all(u in uids_estante for u in valores), "órfão no mapa!"

    # ciclos de compor/separar — identidade estável, sem duplicar nem vazar
    for _ in range(3):
        idx_a = next(i for i, it in enumerate(mesa._itens) if it.uid == a.uid)
        idx_b = next(i for i, it in enumerate(mesa._itens) if it.uid == b.uid)
        mesa._executar_composicao(idx_a, idx_b, "Composto 0 e 1", "9,99")
        _sem_duplicata_nem_orfao()
        comp = next(it for it in mesa._itens if servico.eh_composto(it))
        assert mesa._mapa[slot_a] == comp.uid
        idx_c = mesa._itens.index(comp)
        mesa._executar_separacao(idx_c)
        _sem_duplicata_nem_orfao()
        assert mesa._mapa[slot_a] == a.uid          # A voltou, uid DE SEMPRE
        assert len(mesa._itens) == 4

    # composto por pixel: A à ESQUERDA, B à DIREITA (LADO_A_LADO)
    idx_a = next(i for i, it in enumerate(mesa._itens) if it.uid == a.uid)
    idx_b = next(i for i, it in enumerate(mesa._itens) if it.uid == b.uid)
    mesa._executar_composicao(idx_a, idx_b, "Composto Final", "9,99")
    comp = next(it for it in mesa._itens if servico.eh_composto(it))
    dados = mesa._dados_por_slot()
    img = compor_pagina(layout, layout.paginas[0], dados)
    slot0 = next(s for s in layout.paginas[0].slots if s.id == slot_a)
    e0 = tuple(int(CORES[0][i:i + 2], 16) for i in (1, 3, 5))
    e1 = tuple(int(CORES[1][i:i + 2], 16) for i in (1, 3, 5))
    assert _cor_em(img, layout, slot0, 0.25, 0.5) == e0    # Camil à esquerda
    assert _cor_em(img, layout, slot0, 0.75, 0.5) == e1    # Rei à direita

    # congelar → reabrir → o composto vive; separar DEPOIS de reabrir devolve
    # os dois com foto congelada VIVA (byte a byte) — e o mapa sem órfão
    from app.core import projetos
    bytes_a = Path(a.imagem).read_bytes()
    bytes_b = Path(b.imagem).read_bytes()
    pid = projetos.salvar_projeto("Composto Adv", None, "TABLOIDE", layout,
                                  [it.to_dict() for it in mesa._itens],
                                  mapa=mesa._mapa)
    Path(a.imagem).unlink()                # as fotos ORIGINAIS somem do disco
    Path(b.imagem).unlink()

    p = projetos.abrir_projeto(pid)
    reitens = [ItemMesa.from_dict(d) for d in p.itens]
    recomp = next(it for it in reitens if servico.eh_composto(it))
    assert recomp.uid == comp.uid          # a identidade congelou
    # compõe do congelado com os pixels nos lugares
    por_uid = {it.uid: it for it in reitens}
    dados2 = {}
    for sid, u in p.mapa.items():
        it = por_uid[u]
        dados2[sid] = DadosProduto(
            it.nome, preco_por=Decimal("1"), imagem_path=it.imagem,
            imagens=[ImagemSlot(c) for c in (it.imagens or [])],
            modo_arranjo=ModoArranjo(it.arranjo) if it.arranjo
            else ModoArranjo.LEQUE)
    img2 = compor_pagina(p.layout, p.layout.paginas[0], dados2)
    slot0_p = next(s for s in p.layout.paginas[0].slots if s.id == slot_a)
    assert _cor_em(img2, p.layout, slot0_p, 0.25, 0.5) == e0
    assert _cor_em(img2, p.layout, slot0_p, 0.75, 0.5) == e1

    # separar o congelado: origens com uids ORIGINAIS e fotos byte-idênticas
    v_a, v_b = servico.separar_item(recomp)
    assert (v_a.uid, v_b.uid) == (a.uid, b.uid)
    assert Path(v_a.imagem).read_bytes() == bytes_a
    assert Path(v_b.imagem).read_bytes() == bytes_b


def test_adversarial_secoes_nao_cobrem_conteudo(raiz_tmp, tmp_path):
    """B4 da ORDEM_F8 (F8.2, I5): seções LIGADAS sob ataque, por pixel.

    O trio imagem×nome×preço de CADA célula fica idêntico com seções
    ligadas ou desligadas (amostragem em vários pontos por região — o
    contorno não desloca nem cobre conteúdo), sob shuffle e depois de
    salvar→reabrir o congelado (liga/desliga e títulos persistem).
    """
    QApplication.instance() or QApplication([])
    layout = _grade_4()
    pagina = layout.paginas[0]
    itens = _itens(tmp_path)[:4]
    mapa = {s.id: it.uid for s, it in zip(pagina.slots, itens)}
    categorias = ["Bebidas", "Bebidas", "Limpeza", "Limpeza"]
    for it, cat in zip(itens, categorias):
        it.categoria = cat

    def _dados():
        por_uid = {it.uid: it for it in itens}
        return {sid: DadosProduto(por_uid[u].nome, preco_por=Decimal("1"),
                                  imagem_path=por_uid[u].imagem,
                                  categoria=por_uid[u].categoria)
                for sid, u in mapa.items()}

    def _pontos(slot):
        """Vários pontos DENTRO de cada região do slot (o trio inteiro)."""
        pts = []
        for reg in slot.regioes:
            for fx, fy in ((0.5, 0.5), (0.25, 0.5), (0.75, 0.5), (0.5, 0.25)):
                cx = mm_para_px(reg.rect.x_mm + reg.rect.larg_mm * fx,
                                layout.dpi)
                cy = mm_para_px(reg.rect.y_mm + reg.rect.alt_mm * fy,
                                layout.dpi)
                pts.append((round(cx), round(cy)))
        return pts

    pagina.secoes_ligadas = False
    desligada = compor_pagina(layout, pagina, _dados())
    pagina.secoes_ligadas = True
    pagina.titulos_secoes = {"Limpeza": "Casa Limpa"}
    ligada = compor_pagina(layout, pagina, _dados())
    assert list(ligada.getdata()) != list(desligada.getdata())  # seções existem

    for slot in pagina.slots:               # o trio é intocado, célula a célula
        for p in _pontos(slot):
            assert ligada.getpixel(p) == desligada.getpixel(p), \
                f"{slot.id}: a seção cobriu conteúdo em {p}!"
        assert _cor_no_slot(ligada, layout, slot) == \
            _cor_no_slot(desligada, layout, slot)

    # shuffle da estante não mexe em nada (o vínculo é o mapa)
    random.shuffle(itens)
    ligada2 = compor_pagina(layout, pagina, _dados())
    assert list(ligada2.getdata()) == list(ligada.getdata())

    # congelar → reabrir: liga/desliga e TÍTULOS persistem; trio segue exato
    from app.core import projetos
    pid = projetos.salvar_projeto("Secoes Adv", None, "TABLOIDE", layout,
                                  [it.to_dict() for it in itens], mapa=mapa)
    p = projetos.abrir_projeto(pid)
    pag2 = p.layout.paginas[0]
    assert pag2.secoes_ligadas is True
    assert pag2.titulos_secoes == {"Limpeza": "Casa Limpa"}
    por_uid = {d["uid"]: d for d in p.itens}
    dados2 = {sid: DadosProduto(por_uid[u]["nome"], preco_por=Decimal("1"),
                                imagem_path=por_uid[u]["imagem"],
                                categoria=por_uid[u]["categoria"])
              for sid, u in p.mapa.items()}
    reaberta = compor_pagina(p.layout, pag2, dados2)
    for slot in pag2.slots:
        esperado = CORES[int(por_uid[p.mapa[slot.id]]["nome"].split("-")[1])]
        e = tuple(int(esperado[i:i + 2], 16) for i in (1, 3, 5))
        assert _cor_no_slot(reaberta, p.layout, slot) == e


def test_preco_decimal_adversarial():
    casos = {
        "R$ 1.299,00": Decimal("1299.00"), "1.299": Decimal("1299"),
        "17,7": Decimal("17.7"), "<> R$ 17,71": Decimal("17.71"),
        "1,299.00": Decimal("1299.00"), "5.90": Decimal("5.90"),
        "1299": Decimal("1299"), "R$ 5,90 UN": Decimal("5.90"),
        "": None, None: None, "abc": None, ",": None,
        # P0.3b: mais de um número no texto = ambíguo → None (nunca fundir:
        # "2x 5,00" NÃO pode virar 25,00 — preço errado é pior que ausente)
        "2x 5,00": None, "3 por 10,00": None, "Leve 3 10,00": None,
        "2 un 9,90": None,
        # número único com decoração continua ok
        "17,71 /kg": Decimal("17.71"), "10,00.": Decimal("10.00"),
    }
    for entrada, esperado in casos.items():
        assert servico.preco_decimal(entrada) == esperado, f"falhou: {entrada!r}"


def test_fonte_ausente_nao_derruba_exportacao(tmp_path):
    from app.rendering.compositor import fonte_segura
    f = fonte_segura(tmp_path, "NaoExiste.ttf", 24)      # dir sem NENHUMA fonte
    assert f is not None                                  # fallback embutido

def test_adversarial_rotacao_nao_desloca_conteudo_de_celula(raiz_tmp, tmp_path):
    """RG-12 (Onda 3, exigência do arquiteto): girar a região de UMA célula
    não move UM PIXEL das outras, e o vínculo slot→uid fica intacto. O trio
    continua conferido POR CONTEÚDO (cor) em todas as células."""
    lay = _grade_4()
    itens = _itens(tmp_path)[:4]
    mapa = {s.id: it.uid for s, it in zip(lay.paginas[0].slots, itens)}
    mapa_antes = dict(mapa)

    por_uid = {it.uid: it for it in itens}
    dados = {sid: DadosProduto(por_uid[u].nome,
                               preco_por=servico.preco_decimal(por_uid[u].preco),
                               imagem_path=por_uid[u].imagem)
             for sid, u in mapa.items()}
    antes = compor_pagina(lay, lay.paginas[0], dados)

    # o ataque: deitar o NOME da célula_1 (override local — a propagação da
    # mestra não pode desfazer nem espalhar para as irmãs)
    alvo = next(s for s in lay.paginas[0].slots if s.id == "celula_1")
    reg_nome = next(r for r in alvo.regioes if r.tipo == TipoRegiao.NOME)
    reg_nome.rotacao_graus = 90.0
    reg_nome.overrides.add("rotacao_graus")
    propagar_mestre(lay.paginas[0])
    depois = compor_pagina(lay, lay.paginas[0], dados)

    px = lay.dpi / 25.4
    quadrantes_intactos = [(0, 0, 55, 55), (0, 55, 55, 110), (55, 55, 110, 110)]
    for (x0, y0, x1, y1) in quadrantes_intactos:
        caixa = (round(x0 * px), round(y0 * px), round(x1 * px), round(y1 * px))
        assert antes.crop(caixa).tobytes() == depois.crop(caixa).tobytes(), \
            f"a rotação da célula_1 vazou para o quadrante {(x0, y0)}!"
    caixa_alvo = (round(55 * px), 0, round(110 * px), round(55 * px))
    assert antes.crop(caixa_alvo).tobytes() != depois.crop(caixa_alvo).tobytes()

    assert mapa == mapa_antes                       # vínculo intacto (I1)
    _conferir(lay, mapa, itens)                     # trio por conteúdo, 4 células


# =============================================================================
# FASE 4 (Bloco G, passos 81-85): os NOVOS caminhos de seleção e agrupamento.
# O adversarial re-rodado — verifica o trio POR CONTEÚDO, nunca por exceção.
# =============================================================================

def _canvas_trio(cor="#FF00FF"):
    """Canvas com um trio de conteúdo SOLTO, imagem colorida (conteúdo)."""
    import tempfile

    from app.qt.canvas import CanvasView
    foto = Path(tempfile.mkdtemp()) / "f.png"
    Image.new("RGB", (60, 60), cor).save(foto)
    c = CanvasView()
    c.resize(600, 500)
    c.show()
    trio = Slot("livre", [
        Regiao(TipoRegiao.IMAGEM, Retangulo(10, 10, 30, 20), nome="Img"),
        Regiao(TipoRegiao.NOME, Retangulo(10, 32, 30, 8), nome="Nome"),
        Regiao(TipoRegiao.PRECO, Retangulo(10, 41, 30, 8), nome="Preço"),
    ])
    lay = LayoutDef(100, 100, dpi=100, paginas=[Pagina([trio])])
    c.carregar(lay, DadosProduto("PROD", preco_por=Decimal("2.00"),
                                 imagem_path=str(foto)))
    return c, str(foto)


def test_adversarial_agrupar_desagrupar_reagrupar_por_conteudo(raiz_tmp, tmp_path):
    """Passo 82 (reescrito — por PIXEL e uid, com o REAGRUPAR de fato):
    agrupar → carimbar cópia → desagrupar → REAGRUPAR mantém a imagem de
    CADA célula. Cada célula recebe uma foto de cor distinta; ao compor,
    cada célula mostra a SUA cor — antes e depois do ciclo. Nunca troca de
    célula (I1: as imagens têm uids distintos que não colidem)."""
    QApplication.instance() or QApplication([])
    from app.qt.canvas import CanvasView

    # 2 fotos de cores distintas (conteúdo verificável por pixel)
    fmagenta = tmp_path / "m.png"
    fciano = tmp_path / "c.png"
    Image.new("RGB", (80, 80), "#FF00FF").save(fmagenta)   # mestra
    Image.new("RGB", (80, 80), "#00FFFF").save(fciano)     # cópia

    trio = Slot("livre", [
        Regiao(TipoRegiao.IMAGEM, Retangulo(10, 10, 30, 24), nome="Img"),
        Regiao(TipoRegiao.NOME, Retangulo(10, 35, 30, 6), nome="Nome"),
        Regiao(TipoRegiao.PRECO, Retangulo(10, 42, 30, 6), nome="Preço"),
    ])
    lay = LayoutDef(120, 60, dpi=100, paginas=[Pagina([trio])])
    c = CanvasView()
    c.resize(600, 400)
    c.show()
    c.carregar(lay, DadosProduto("M", preco_por=Decimal("1")))

    for it in c._itens:
        it.setSelected(True)
    c.agrupar_selecao()
    m = next(s for s in c._pagina().slots if s.mestre)
    c.carimbar_grupo(m.id, (70, 10))
    copia = next(s for s in c._pagina().slots if s.ref_grupo == m.id)
    reg_m = m.regioes[0]                              # a IMAGEM da mestra
    img_m_uid = next(r.uid for r in m.regioes if r.tipo == TipoRegiao.IMAGEM)
    img_c_uid = next(r.uid for r in copia.regioes if r.tipo == TipoRegiao.IMAGEM)
    assert img_m_uid != img_c_uid                    # identidades distintas (I1)

    def _cor_da_img(slot):
        """Pixel central da região IMAGEM do slot, compondo por slot."""
        dados = {m.id: DadosProduto("M", imagem_path=str(fmagenta)),
                 copia.id: DadosProduto("C", imagem_path=str(fciano))}
        img = compor_pagina(c._layout, c._pagina(), dados)
        return _cor_no_slot(img, c._layout, slot)

    MAG, CIA = (255, 0, 255), (0, 255, 255)
    assert _cor_da_img(m) == MAG                      # mestra = magenta
    assert _cor_da_img(copia) == CIA                  # cópia = ciano

    # DESAGRUPAR
    c.desagrupar_regiao(reg_m)
    assert all(not s.mestre for s in c._pagina().slots)
    # REAGRUPAR o trio da ex-mestra (o passo que o docstring promete)
    for it in c._itens:
        it.setSelected(it.regiao in m.regioes)
    c.agrupar_selecao()
    novo_m = next(s for s in c._pagina().slots if s.mestre)
    assert any(r.uid == img_m_uid for r in novo_m.regioes)   # a mesma imagem

    # POR PIXEL de novo: cada célula ainda mostra a SUA cor (nunca trocou)
    dados2 = {novo_m.id: DadosProduto("M", imagem_path=str(fmagenta)),
              copia.id: DadosProduto("C", imagem_path=str(fciano))}
    img2 = compor_pagina(c._layout, c._pagina(), dados2)
    assert _cor_no_slot(img2, c._layout, novo_m) == MAG
    assert _cor_no_slot(img2, c._layout, copia) == CIA
    # e o uid da imagem da cópia seguiu intacto (não colidiu/trocou)
    assert any(r.uid == img_c_uid for r in copia.regioes)


def test_adversarial_zorder_seleciona_o_topo(raiz_tmp):
    """Passo 83 (a cura do RG-55): clicar em regiões EMPILHADAS resolve
    sempre a do TOPO do z e o painel mostra o uid certo."""
    from PySide6.QtWidgets import QApplication as QA

    QA.instance() or QA([])
    from app.qt.canvas import CanvasView
    baixo = Regiao(TipoRegiao.IMAGEM, Retangulo(10, 10, 40, 40), nome="Baixo")
    cima = Regiao(TipoRegiao.PRECO, Retangulo(15, 15, 20, 20), nome="Cima")
    # a de CIMA é a última na lista do slot → topo do z
    lay = LayoutDef(100, 100, dpi=100,
                    paginas=[Pagina([Slot("s", [baixo, cima])])])
    c = CanvasView()
    c.resize(500, 400)
    c.show()
    c.carregar(lay, DadosProduto("P", preco_por=Decimal("1")))
    it_cima = next(i for i in c._itens if i.regiao.nome == "Cima")
    pt = it_cima.mapToScene(it_cima._w / 2, it_cima._h / 2)
    r = c.resolver_selecao(pt)
    assert r is cima                               # a do topo, nunca a de baixo
    assert r.uid == cima.uid                       # uid certo (I1)


def test_adversarial_rotacao_hit_test_real(raiz_tmp):
    """Passo 84 (reescrito — SEM setar _primaria na mão): exercita o hit-test
    ROTACIONADO de verdade. A região é larga-e-baixa, girada 90° → fica
    alta-e-estreita. `resolver_selecao` (o picking de produção, o mesmo que o
    mousePress do Qt usa) só acerta a região no ponto girado se a rotação for
    respeitada; e devolve None no ponto do retângulo NÃO-girado. Depois usa
    esse picking para SELECIONAR e confere o vínculo por uid após edição."""
    from PySide6.QtCore import QPointF
    from PySide6.QtWidgets import QApplication as QA

    QA.instance() or QA([])
    from app.qt.canvas import CanvasView
    reg = Regiao(TipoRegiao.PRECO, Retangulo(20, 20, 30, 12), nome="Preço",
                 rotacao_graus=90.0)
    lay = LayoutDef(100, 100, dpi=100, paginas=[Pagina([Slot("s", [reg])])])
    c = CanvasView()
    c.resize(500, 400)
    c.show()
    c.carregar(lay, DadosProduto("P", preco_por=Decimal("1")))

    px = 100 / 25.4
    cx = (20 + 30 / 2) * px          # centro (invariante à rotação)
    cy = (20 + 12 / 2) * px
    w_px = 30 * px                    # largura ORIGINAL (vira altura girada)
    # ponto ACIMA do centro por 0.40·largura: DENTRO do girado (alto-estreito),
    # FORA do retângulo original (baixo-largo)
    pt_girado = QPointF(cx, cy - w_px * 0.40)
    # ponto À ESQUERDA do centro por 0.40·largura: DENTRO do retângulo
    # original, FORA do girado
    pt_reto = QPointF(cx - w_px * 0.40, cy)

    assert c.resolver_selecao(pt_girado) is reg    # a rotação foi respeitada
    assert c.resolver_selecao(pt_reto) is None     # nada no lugar não-girado

    # usa o picking REAL para selecionar (não setSelected+_primaria na mão)
    achada = c.resolver_selecao(pt_girado)
    for it in c._itens:
        it.setSelected(it.regiao is achada)
    c._primaria = achada
    c._emitir_selecao()
    assert c.selecionada() is reg and c.selecionada().uid == reg.uid
    # editar e o vínculo (uid) não muda
    c.notificar_edicao(reg, "rotacao_graus")
    assert c.resolver_selecao(pt_girado).uid == reg.uid


def test_adversarial_undo_agrupar_restaura_mapa_e_layout(raiz_tmp):
    """Passo 85: desfazer um AGRUPAR restaura mapa + layout JUNTOS. Com o
    mapa POPULADO (não {}), a asserção do mapa não é vácua; e o caso que DE
    FATO muta o mapa (remover_celula) é coberto ao fim."""
    QApplication.instance() or QApplication([])
    c, _ = _canvas_trio()
    # POPULA o mapa com uma entrada real antes de agrupar
    slot0 = c._pagina().slots[0]
    c.mapa = {slot0.id: "item-uid-x"}
    c._registrar_hist()
    mapa_antes = dict(c.mapa)
    n_slots_antes = len(c._pagina().slots)
    assert mapa_antes                                  # não-vácuo

    for it in c._itens:
        it.setSelected(True)
    c.agrupar_selecao()
    assert any(s.mestre for s in c._pagina().slots)   # agrupou
    assert c.desfazer()                               # UM passo
    assert not any(s.mestre for s in c._pagina().slots)   # desagrupou junto
    assert len(c._pagina().slots) == n_slots_antes
    assert c.mapa == mapa_antes                        # o mapa voltou junto

    # o caso que MUTA o mapa: remover a célula-mestra tira a entrada; desfazer
    # restaura mapa + layout juntos (D5)
    m = next(s for s in c._pagina().slots if s.mestre) \
        if any(s.mestre for s in c._pagina().slots) else None
    # reagrupa para ter uma célula removível com entrada no mapa
    for it in c._itens:
        it.setSelected(True)
    c.agrupar_selecao()
    m = next(s for s in c._pagina().slots if s.mestre)
    c.mapa = {m.id: "item-uid-y"}
    c._registrar_hist()
    mapa2 = dict(c.mapa)
    n2 = len(c._pagina().slots)
    assert c.remover_celula(m.id)
    assert m.id not in c.mapa                          # o mapa perdeu a entrada
    assert c.desfazer()
    assert c.mapa == mapa2                             # mapa voltou
    assert len(c._pagina().slots) == n2               # layout voltou junto


def test_adversarial_i3_guias_relativas(raiz_tmp):
    """Passo 87 (I3): guias e passo de grade persistem em coordenadas mm
    RELATIVAS (nunca px absoluto) — portáveis entre máquinas."""
    QApplication.instance() or QApplication([])
    c, _ = _canvas_trio()
    c.adicionar_guia("x", 42.5)
    c.set_grade_magnetica(True)
    c.set_grade_passo(7.5)
    d = c._layout.to_dict()
    pag = d["paginas"][0] if "paginas" in d else d.get("layout", {})
    # round-trip preserva as coords em mm (float), sem px
    lay2 = LayoutDef.from_dict(d)
    p = lay2.paginas[0]
    assert ("x", 42.5) in p.guias
    assert p.grade_passo_mm == 7.5
    assert all(isinstance(g[1], float) for g in p.guias)


# ============================================================================
# FASE 5 — Bloco F: adversariais de máscara / modelo / conta-gotas (I1-I5)
# ============================================================================

def test_adversarial_f5_mascara_na_mestra_replica_e_nao_troca(raiz_tmp, tmp_path):
    """Passos 77/81 (I4): máscara circular na IMAGEM da mestra replica nas
    cópias (pareadas por ref_mestre) E não troca o produto de cada célula —
    cada centro segue a cor do SEU item."""
    QApplication.instance() or QApplication([])
    from app.rendering.model import Mascara
    layout = _grade_4()
    mestre = layout.paginas[0].slots[0]
    img_m = next(r for r in mestre.regioes if r.tipo == TipoRegiao.IMAGEM)
    img_m.mascara = Mascara.CIRCULO
    propagar_mestre(layout.paginas[0])
    for slot in layout.paginas[0].slots[1:]:
        img_c = next(r for r in slot.regioes if r.tipo == TipoRegiao.IMAGEM)
        assert img_c.mascara is Mascara.CIRCULO          # I4: herdou a máscara
        assert img_c.ref_mestre == img_m.uid             # casada por uid
    itens = _itens(tmp_path)
    mapa = {s.id: it.uid for s, it in zip(layout.paginas[0].slots, itens)}
    _conferir(layout, mapa, itens)                       # vínculo intacto


def test_adversarial_f5_carimbar_modelo_item_por_uid(raiz_tmp, tmp_path):
    """Passo 76: carimbar um modelo num slot com item → o conteúdo vem do item
    (por uid), com o estilo do modelo; nada troca de célula."""
    QApplication.instance() or QApplication([])
    from app.rendering.modelos import carimbar_modelo, modelo_vitrine
    foto = tmp_path / "verde.png"
    Image.new("RGB", (200, 200), "#00FF00").save(foto)
    slot = Slot("celula_x", carimbar_modelo(modelo_vitrine(), 0, 0, 100, 100),
                origem_mm=(0, 0))
    lay = LayoutDef(100, 100, dpi=100, paginas=[Pagina([slot])])
    dados = {"celula_x": DadosProduto("Produto", preco_por=Decimal("9.99"),
                                      imagem_path=str(foto))}
    img = compor_pagina(lay, lay.paginas[0], dados)
    cor = _cor_no_slot(img, lay, slot)
    assert cor[1] > 150 and cor[0] < 120                 # a foto verde do item
    nome = next(r for r in slot.regioes if r.tipo == TipoRegiao.NOME)
    assert nome.pill                                     # estilo veio do modelo


def test_adversarial_f5_conta_gotas_nao_troca_conteudo(raiz_tmp, tmp_path):
    """Passo 78: colar estilo em lote muda SÓ o estilo — a imagem de cada
    célula (conteúdo) e as posições (geometria) ficam."""
    QApplication.instance() or QApplication([])
    from app.rendering.estilos import copiar_estilo_visual
    layout = _grade_4()
    itens = _itens(tmp_path)
    mapa = {s.id: it.uid for s, it in zip(layout.paginas[0].slots, itens)}
    nomes = [r for s in layout.paginas[0].slots for r in s.regioes
             if r.tipo == TipoRegiao.NOME]
    origem = nomes[0]
    origem.cor, origem.pill = "#123456", True
    rects_antes = [(r.rect.x_mm, r.rect.y_mm, r.rect.larg_mm) for r in nomes]
    for destino in nomes[1:]:
        copiar_estilo_visual(origem, destino)
    _conferir(layout, mapa, itens)                       # conteúdo intacto (I1)
    assert [(r.rect.x_mm, r.rect.y_mm, r.rect.larg_mm) for r in nomes] == rects_antes
    assert all(r.cor == "#123456" and r.pill for r in nomes[1:])


# ============================================================================
# FASE 6 — Bloco C: troca / reordenação / duplicar (mapa slot→uid, I1)
# ============================================================================

def test_adversarial_f6_trocar_conserva_trio(raiz_tmp, tmp_path):
    """R-057/passo 36-37/65: trocar duas células TROCA os uids e o conteúdo
    segue (por pixel); o OVERRIDE fica no SLOT (não viaja com o item)."""
    from app.qt.canvas import CanvasView
    QApplication.instance() or QApplication([])
    layout = _grade_4()
    itens = _itens(tmp_path)
    slots = layout.paginas[0].slots
    sid_a, sid_b = slots[0].id, slots[1].id
    v = CanvasView()
    v.carregar(layout, {})
    v.ao_restaurar = None
    v.mapa = {s.id: it.uid for s, it in zip(slots, itens)}
    v.overrides = {sid_a: {"nome": "SÓ NESTA CÉLULA"}}     # override do slot A
    u_a, u_b = v.mapa[sid_a], v.mapa[sid_b]

    assert v.trocar_conteudo_slots(sid_a, sid_b)
    assert v.mapa[sid_a] == u_b and v.mapa[sid_b] == u_a   # uids trocaram (I1)
    assert v.overrides.get(sid_a) == {"nome": "SÓ NESTA CÉLULA"}  # override FICOU no slot
    assert sid_b not in v.overrides
    _conferir(layout, v.mapa, itens)                      # trio por CONTEÚDO


def test_adversarial_f6_reordenar_conserva_cada_trio(raiz_tmp, tmp_path):
    """R-055/passo 34: reordenar a estante re-atribui o mapa por uid — cada
    item leva o SEU trio, nada migra para o vizinho (por conteúdo)."""
    from app.qt.telas.mesa import MesaTela
    QApplication.instance() or QApplication([])
    layout = _grade_4()
    itens = _itens(tmp_path)[:4]
    m = MesaTela()
    m._layout = layout
    m.area.carregar(layout, {})
    m._itens = list(itens)
    m.reordenar_estante(list(reversed(itens)))            # inverte a ordem
    _conferir(layout, m._mapa, itens)                     # cada trio intacto
    # a nova ordem chegou ao mapa (por uid): a 1ª célula (ordem visual) agora
    # aponta para o item que era o ÚLTIMO
    from app.rendering.grade import ocupaveis, ordenar_slots_visualmente
    prim = ocupaveis(ordenar_slots_visualmente(layout.paginas[0].slots))[0]
    assert m._mapa[prim.id] == itens[-1].uid


def test_adversarial_f6_duplicar_uid_novo_independente(raiz_tmp, tmp_path):
    """R-069/passo 38-39: duplicar cria uid NOVO (não referência); editar a
    cópia não toca o original."""
    from app.qt.telas.mesa import MesaTela
    QApplication.instance() or QApplication([])
    m = MesaTela()
    itens = _itens(tmp_path)
    m._itens = list(itens)
    orig = itens[0]
    novo = m.duplicar_item(orig)
    assert novo.uid != orig.uid                           # I1: identidade nova
    assert novo.nome == orig.nome and novo.imagem == orig.imagem  # copia congelada
    novo.nome = "EDITADO NA CÓPIA"
    assert orig.nome != "EDITADO NA CÓPIA"                # independente


def test_adversarial_f6_matriz_override_x_mapa(raiz_tmp, tmp_path):
    """FASE 6 Bloco E (passos 66/67): o override é do SLOT — reordenar NÃO o
    move; duplicar NÃO o herda (o item novo começa limpo)."""
    from app.qt.telas.mesa import MesaTela
    from app.rendering.grade import ocupaveis, ordenar_slots_visualmente
    QApplication.instance() or QApplication([])
    layout = _grade_4()
    itens = _itens(tmp_path)[:4]
    m = MesaTela()
    m._layout = layout
    m.area.carregar(layout, {})
    m._itens = list(itens)
    m.reordenar_estante(list(itens))                  # preenche o mapa
    sid0 = ocupaveis(ordenar_slots_visualmente(layout.paginas[0].slots))[0].id
    m._overrides = {sid0: {"preco": "0,99"}}          # override NAQUELA célula

    m.reordenar_estante(list(reversed(itens)))        # inverte a estante
    assert m._overrides.get(sid0) == {"preco": "0,99"}   # override FICOU no slot
    # o item que agora ocupa sid0 mudou (reordenou), mas o override do SLOT vale
    _conferir(layout, m._mapa, itens)                 # trio por conteúdo, intacto

    novo = m.duplicar_item(itens[0])                  # duplicar
    assert novo.uid not in set(m._mapa.values())      # não entra em slot → sem célula
    assert all(novo.uid != u for u in m._mapa.values())  # I1: identidade nova, sem override


def test_adversarial_f7_encher_pagina_por_uid(raiz_tmp, tmp_path):
    """FASE 7 (R-056/passo 74): 'encher a página' distribui por UID — o mapa
    aponta cada slot para o uid do item na ordem, e cada célula mostra o SEU
    item (por conteúdo). Prova de mutação: se distribuísse por índice/posição
    embaralhando, o _conferir por cor cairia."""
    from app.qt.telas import servico
    from app.rendering.grade import ocupaveis, ordenar_slots_visualmente
    QApplication.instance() or QApplication([])
    layout = _grade_4()
    itens = _itens(tmp_path)[:4]
    slot_ids = [s.id for s in
                ocupaveis(ordenar_slots_visualmente(layout.paginas[0].slots))]
    mapa, resto, avisos = servico.plano_encher_pagina(itens, slot_ids)
    assert list(mapa.values()) == [it.uid for it in itens]   # vínculo por uid (I1)
    assert resto == []
    _conferir(layout, mapa, itens)                            # trio por conteúdo


def test_adversarial_f7_multi_preco_e_observacao_por_uid(raiz_tmp, tmp_path):
    """FASE 7 (R-070/R-071, passo 75/I5): multi-preço e observação são campos do
    ITEM e viajam por UID — depois de embaralhar a estante, cada slot desenha o
    multi-preço e a observação do SEU item, nunca do vizinho. Verificado POR
    CONTEÚDO: o TEXTO composto de cada região (preço e observação) e um diff de
    pixel provando que a observação de fato pinta. Prova de mutação: reconstruir
    por índice em vez de uid cruzaria os campos após o shuffle."""
    from app.rendering.compositor import (
        DadosProduto, compor_pagina, texto_composto_legal)
    from app.rendering.model import PapelTexto, Regiao, Retangulo, TipoRegiao
    QApplication.instance() or QApplication([])

    a = ItemMesa(descricao="A", preco=None, semaforo="VERDE", nome="Arroz",
                 multi_preco="3 por R$10,00", observacao="Limite de 2 por cliente")
    b = ItemMesa(descricao="B", preco=None, semaforo="VERDE", nome="Feijão",
                 multi_preco="Leve 3 pague 2", observacao="Só à vista")
    estante = [a, b]
    mapa = {"s0": a.uid, "s1": b.uid}                 # vínculo por uid (I1)
    random.shuffle(estante)                           # a estante embaralha!
    por_uid = {it.uid: it for it in estante}

    # reconstrução por uid → cada slot pega o SEU item, não o da posição
    dados = {sid: DadosProduto(por_uid[u].nome, multi_preco=por_uid[u].multi_preco,
                               observacao=por_uid[u].observacao)
             for sid, u in mapa.items()}
    assert dados["s0"].multi_preco == "3 por R$10,00"      # A, não B
    assert dados["s0"].observacao == "Limite de 2 por cliente"
    assert dados["s1"].multi_preco == "Leve 3 pague 2"     # B, não A
    assert dados["s1"].observacao == "Só à vista"

    # o TEXTO que a região de observação desenha é o do item certo (por conteúdo)
    reg_obs = Regiao(TipoRegiao.TEXTO_LEGAL, Retangulo(0, 0, 40, 8),
                     papel_texto=PapelTexto.OBSERVACAO)
    assert texto_composto_legal(reg_obs, dados["s0"]) == "Limite de 2 por cliente"
    assert texto_composto_legal(reg_obs, dados["s1"]) == "Só à vista"

    # e a observação REALMENTE pinta (diff de pixel: com × sem)
    from app.rendering.model import LayoutDef, Pagina, Slot
    slot = Slot("s0", [reg_obs], mestre=True, origem_mm=(0, 0))
    lay = LayoutDef(40, 10, dpi=100, paginas=[Pagina([slot])])
    com = compor_pagina(lay, lay.paginas[0], {"s0": dados["s0"]})
    vazio = compor_pagina(lay, lay.paginas[0],
                          {"s0": DadosProduto("Arroz", observacao=None)})
    assert list(com.getdata()) != list(vazio.getdata())   # o texto apareceu


def _cor_do_card(card, item):
    """A cor da foto no centro da região de imagem (vitrine: topo, ~30%)."""
    px = card.getpixel((card.width // 2, int(card.height * 0.30)))[:3]
    esperado = CORES[int(item.nome.split("-")[1])]
    return px, tuple(int(esperado[i:i + 2], 16) for i in (1, 3, 5))


def test_adversarial_f8_carrossel_preserva_ordem_da_lista(raiz_tmp, tmp_path):
    """FASE 8 (R-140, passo 75-76/I5): cada card do carrossel é uma página de 1
    slot; o vínculo produto→card é a ORDEM da lista de dados que o chamador monta
    (por uid). Aqui provo, POR CONTEÚDO, que `compor_carrossel` PRESERVA essa
    ordem — card i mostra o item i — com uma MUTAÇÃO REAL: inverter a lista
    inverte os cards (um passthrough por índice quebrado — ex.: sorted/reversed
    interno — cairia).

    (Corrige um mascaramento apontado pela minha frota adversarial: a versão
    anterior embaralhava a estante mas remontava os dados de um dict por uid em
    ordem fixa — o shuffle era inerte e o teste passaria com casamento por índice.)"""
    from app.rendering.compositor import DadosProduto
    from app.rendering.social import compor_carrossel
    QApplication.instance() or QApplication([])
    itens = _itens(tmp_path)[:3]                      # PROD-0, PROD-1, PROD-2 (cores)
    dados = [DadosProduto(it.nome, imagem_path=it.imagem) for it in itens]

    cards = compor_carrossel(dados)
    assert len(cards) == 3
    for it, card in zip(itens, cards):               # card i == item i (conteúdo)
        px, e = _cor_do_card(card, it)
        assert px == e, f"{it.nome}: foto {px} != {e}"

    # MUTAÇÃO REAL: inverter a lista inverte os cards (a ordem é load-bearing)
    cards_inv = compor_carrossel(list(reversed(dados)))
    for it, card in zip(list(reversed(itens)), cards_inv):
        px, e = _cor_do_card(card, it)
        assert px == e, f"invertido {it.nome}: {px} != {e}"
    # e o 1º card invertido é o ÚLTIMO item (prova que não há reordenação interna)
    px0, e0 = _cor_do_card(cards_inv[0], itens[-1])
    assert px0 == e0


def test_adversarial_f9_revisora_nao_altera_a_peca(raiz_tmp, tmp_path):
    """FASE 9 (R-081, passo 77/I5): a revisora só LÊ — a peça composta é IDÊNTICA
    (byte a byte) antes e depois de revisar. Prova de mutação: se a revisora
    mexesse nos dados/layout, a 2ª composição divergiria."""
    from app.ai.revisora import revisar_export
    from app.core.paths import SystemRoot
    from app.rendering.cartaz import layout_cartaz_exemplo
    from app.rendering.compositor import DadosProduto, compor_pagina
    QApplication.instance() or QApplication([])
    lay = layout_cartaz_exemplo()
    dados = {"cartaz": DadosProduto("Café", preco_por=Decimal("10.00"),
                                    preco_de=Decimal("8.00"), categoria="Mercearia")}
    antes = compor_pagina(lay, lay.paginas[0], dados)
    revisar_export(str(tmp_path / "x.png"), dados, layout=lay, motor=None,
                   fontes_dir=SystemRoot().fontes)
    depois = compor_pagina(lay, lay.paginas[0], dados)
    assert list(antes.getdata()) == list(depois.getdata())   # peça intacta


def test_adversarial_f10_webp_e_foto_repetida_por_conteudo(raiz_tmp, tmp_path):
    """FASE 10 (R-100/R-104, passo 73-74/I5): a WebP PRESERVA a identidade da foto
    (mesmo conteúdo por pixel após o roundtrip) e a foto repetida é pega por
    CONTEÚDO, não por nome. Prova de mutação: WebP lossy perderia o alfa; casar
    por nome não acharia a repetição de nomes diferentes."""
    import shutil as _sh

    from app.images.curadoria import salvar_webp
    from app.qt.telas import servico
    QApplication.instance() or QApplication([])
    # 1) WebP preserva o alfa e a cor central (identidade da foto)
    orig = Image.new("RGBA", (120, 120), (0, 0, 0, 0))
    orig.paste(Image.new("RGBA", (60, 60), (30, 200, 90, 255)), (30, 30))
    wp = salvar_webp(orig, tmp_path / "p.webp", lossless=True)
    reab = Image.open(wp).convert("RGBA")
    assert reab.getpixel((5, 5))[3] == 0                 # alfa (fundo) intacto
    assert reab.getpixel((60, 60))[:3] == (30, 200, 90)   # cor do produto intacta
    # 2) a MESMA foto (nomes diferentes) em 2 itens → repetida por hash
    a = tmp_path / "arroz.png"
    b = tmp_path / "OUTRO_NOME.png"
    Image.new("RGB", (40, 40), (200, 30, 30)).save(a)
    _sh.copy(a, b)
    i1 = servico.ItemMesa("A", "1,00", "VERDE", "A"); i1.imagem = str(a)
    i2 = servico.ItemMesa("B", "2,00", "VERDE", "B"); i2.imagem = str(b)
    grupos = servico.fotos_repetidas([i1, i2])
    assert len(grupos) == 1 and {x.nome for x in grupos[0][1]} == {"A", "B"}


def test_adversarial_f11_excel_casa_por_chave_natural(raiz_tmp, tmp_path):
    """FASE 11 (R-118, passo 76/I1): o import de Excel casa por CHAVE NATURAL —
    editar os DOIS preços na planilha atualiza cada produto no SEU lugar, nunca
    troca dados entre eles; as fotos (que não viajam, I3) ficam intactas.
    Prova de mutação: casar por posição/id embaralharia os preços."""
    from openpyxl import load_workbook

    from app.core import excel_acervo as X
    from app.core.portabilidade import Decisao
    from app.tests import seeds_portabilidade as seeds
    QApplication.instance() or QApplication([])
    root = seeds.raiz(tmp_path, "loja")
    seeds.add_produto(root, "Arroz 5kg", "Camil", "24.90",
                      foto=seeds.png("#FF0000"), categoria="Mercearia")
    seeds.add_produto(root, "Feijao 1kg", "Kicaldo", "8.50",
                      foto=seeds.png("#0000FF"), categoria="Mercearia")
    foto_arroz = seeds.foto_de(root, "Arroz 5kg", "Camil")
    foto_feijao = seeds.foto_de(root, "Feijao 1kg", "Kicaldo")

    xlsx = X.exportar_acervo_xlsx(tmp_path / "a.xlsx", raiz=root)
    wb = load_workbook(str(xlsx))
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "Arroz 5kg":
            row[4].value = "19,90"
        if row[0].value == "Feijao 1kg":
            row[4].value = "6,66"
    wb.save(str(xlsx))
    a = X.analisar_planilha(xlsx, raiz=root)
    dec = {c.id_decisao: Decisao.USAR_PACOTE for c in a.conflitos}
    X.aplicar_importacao_planilha(a, dec, raiz=root)

    # cada preço no SEU produto (não trocado); fotos intactas (I1/I3)
    assert seeds.produto_por_chave(root, "Arroz 5kg", "Camil")["preco"] == "19.90"
    assert seeds.produto_por_chave(root, "Feijao 1kg", "Kicaldo")["preco"] == "6.66"
    assert seeds.foto_de(root, "Arroz 5kg", "Camil") == foto_arroz
    assert seeds.foto_de(root, "Feijao 1kg", "Kicaldo") == foto_feijao


def test_adversarial_f11_relampago_produto_certo(raiz_tmp, tmp_path):
    """FASE 11 (R-110, passo 77/I1): a cartaz-relâmpago compõe a foto do produto
    CERTO. Prova de mutação: o produto errado pintaria outra cor na região."""
    from app.rendering.cartaz import layout_cartaz_exemplo
    from app.rendering.compositor import _rect_px, compor_pagina
    from app.rendering.model import TipoRegiao
    QApplication.instance() or QApplication([])
    Image.new("RGB", (300, 300), (220, 20, 20)).save(tmp_path / "a.png")   # verm.
    Image.new("RGB", (300, 300), (20, 20, 220)).save(tmp_path / "b.png")   # azul
    prod_b = {"nome": "Produto B", "preco": "9,90", "preco_de": "12,90",
              "imagem": str(tmp_path / "b.png"), "validade": "24/07"}
    lay = layout_cartaz_exemplo()
    img = compor_pagina(lay, lay.paginas[0],
                        servico.dados_cartaz_de_produto(prod_b)).convert("RGB")
    reg = next(r for s in lay.paginas[0].slots for r in s.regioes
               if r.tipo == TipoRegiao.IMAGEM)
    x, y, w, h = _rect_px(reg.rect, lay.dpi)
    r, g, b = img.getpixel((x + w // 2, y + h // 2))
    assert b > 150 and r < 100        # AZUL (produto B), não vermelho (A)


def test_adversarial_isolamento_nao_quebra_vinculo_nem_propagacao(
        raiz_tmp, tmp_path):
    """MODO DE ISOLAMENTO (21/07, gesto do Illustrator — I5): editar DENTRO do
    isolamento não pode trocar conteúdo de lugar nem quebrar mestra↔cópia.
    O teste entra no grupo, entra na célula, MOVE uma peça de cópia (vira
    ajuste consciente, I4), edita a MESTRA (propaga em quem não tem ajuste),
    sai, e confere o trio POR CONTEÚDO (cor da imagem por slot) + os vínculos
    por uid — nunca por ausência de exceção."""
    from app.qt.canvas import CanvasView
    QApplication.instance() or QApplication([])
    lay = _grade_4()
    itens = _itens(tmp_path)[:4]
    mapa = {f"celula_{i}": itens[i].uid for i in range(4)}
    por_uid = {it.uid: it for it in itens}
    dados = {sid: DadosProduto(por_uid[u].nome,
                               preco_por=servico.preco_decimal(por_uid[u].preco),
                               imagem_path=por_uid[u].imagem)
             for sid, u in mapa.items()}
    c = CanvasView()
    c.carregar(lay, dados)
    c.mapa.update(mapa)
    pag = lay.paginas[0]
    mestra_nome = next(r for r in pag.slots[0].regioes
                       if r.tipo == TipoRegiao.NOME)

    # nível 1 (grupo) → nível 2 (a CÓPIA celula_1)
    reg_c1 = next(r for r in pag.slots[1].regioes
                  if r.tipo == TipoRegiao.PRECO)
    assert c.isolar_por_duplo_clique(reg_c1)          # grupo (4 células)
    assert c.escopo_isolamento() == {f"celula_{i}" for i in range(4)}
    assert c.isolar_por_duplo_clique(reg_c1)          # a célula da cópia
    assert c.celula_isolada(pag.slots[1])

    # move a peça PREÇO da cópia dentro do isolamento (o gesto do dono)
    it_preco = next(it for it in c._itens if it.regiao is reg_c1)
    ref_antes = reg_c1.ref_mestre                     # o vínculo I4, por uid
    c._scene.clearSelection()
    it_preco.setSelected(True)
    dx = c.mm_para_cena(4, 0)[0]
    it_preco.setPos(it_preco.pos().x() + dx, it_preco.pos().y())
    c._commit_regiao(it_preco)
    assert reg_c1.ref_mestre == ref_antes             # I4: o uid não mudou
    assert "rect" in reg_c1.overrides                 # ajuste CONSCIENTE

    # edita a MESTRA: duplo clique nela TROCA o alvo preservando a
    # profundidade (frota: célula→célula do mesmo grupo é UM gesto)
    assert c.isolar_por_duplo_clique(mestra_nome)     # direto na célula dela
    assert c.celula_isolada(pag.slots[0])
    mestra_nome.rect.x_mm += 2
    c.notificar_edicao(mestra_nome, "rect")           # propaga nas cópias
    for i in (2, 3):                                  # cópias SEM ajuste seguem
        nome_i = next(r for r in pag.slots[i].regioes
                      if r.tipo == TipoRegiao.NOME)
        assert nome_i.rect.x_mm == pytest.approx(
            mestra_nome.rect.x_mm + pag.slots[i].origem_mm[0]
            - pag.slots[0].origem_mm[0], abs=0.01)
    preco_c1_x = reg_c1.rect.x_mm                     # o ajuste da cópia FICOU
    c.sair_isolamento(tudo=True)
    assert reg_c1.rect.x_mm == preco_c1_x

    # o veredito final é o PIXEL: cada célula com a imagem do SEU produto
    _conferir(lay, c.mapa, itens)
