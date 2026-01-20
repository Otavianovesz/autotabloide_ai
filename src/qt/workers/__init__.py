"""
AutoTabloide AI - QThread Workers
===================================
Workers para processamento em background.
"""

from PySide6.QtCore import QThread, Signal, Slot, QMutex, QWaitCondition
from typing import Optional, Callable, Any, List, Dict
from pathlib import Path
import time


class BaseWorker(QThread):
    """Worker base com sinais comuns."""
    
    progress = Signal(int, str)  # valor, mensagem
    status = Signal(str)
    error = Signal(str)
    finished_work = Signal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._running = True
        self._paused = False
        self._mutex = QMutex()
        self._condition = QWaitCondition()
    
    def stop(self) -> None:
        """Para o worker."""
        self._running = False
        if self._paused:
            self.resume()
        self.wait()
    
    def pause(self) -> None:
        """Pausa o worker."""
        self._mutex.lock()
        self._paused = True
        self._mutex.unlock()
    
    def resume(self) -> None:
        """Resume o worker."""
        self._mutex.lock()
        self._paused = False
        self._condition.wakeAll()
        self._mutex.unlock()
    
    def check_pause(self) -> None:
        """Verifica se deve pausar."""
        self._mutex.lock()
        while self._paused:
            self._condition.wait(self._mutex)
        self._mutex.unlock()


class SentinelWorker(BaseWorker):
    """Worker do Sentinel - carrega LLM e processa em background."""
    
    llm_loaded = Signal(bool)  # sucesso
    llm_response = Signal(str, str)  # prompt, resposta
    
    def __init__(self, model_path: str = "", parent=None):
        super().__init__(parent)
        self.model_path = model_path
        self._llm = None
        self._request_queue: List[Dict] = []
    
    def run(self) -> None:
        """Loop principal do Sentinel."""
        self.status.emit("Iniciando Sentinel...")
        
        # Carrega LLM
        if self.model_path:
            self._load_llm()
        
        # Loop de processamento
        while self._running:
            self.check_pause()
            
            if self._request_queue:
                request = self._request_queue.pop(0)
                self._process_request(request)
            else:
                time.sleep(0.1)
        
        self.status.emit("Sentinel finalizado")
        self.finished_work.emit()
    
    def _load_llm(self) -> None:
        """Carrega modelo LLM com fallback graceful."""
        try:
            self.progress.emit(10, "Verificando modelo...")
            
            model_path = Path(self.model_path)
            if not model_path.exists():
                self.status.emit("Modelo LLM não encontrado - modo offline ativado")
                self.llm_loaded.emit(False)
                return
            
            self.progress.emit(50, "Carregando modelo LLM...")
            
            # Tenta carregar llama-cpp-python
            try:
                from llama_cpp import Llama
                
                # Obtém configurações de GPU
                try:
                    from src.core.settings_service import settings_service
                    n_gpu_layers = settings_service.get("llm.n_gpu_layers", 0)
                    n_ctx = settings_service.get("llm.context_size", 2048)
                except Exception:
                    n_gpu_layers = 0
                    n_ctx = 2048
                
                self._llm = Llama(
                    model_path=str(model_path),
                    n_gpu_layers=n_gpu_layers,
                    n_ctx=n_ctx,
                    verbose=False
                )
                
                self.progress.emit(100, "LLM carregado!")
                self.llm_loaded.emit(True)
                self.status.emit("LLM pronto para uso")
                
            except ImportError:
                self.status.emit("llama-cpp-python não instalado - modo offline")
                self.llm_loaded.emit(False)
                
        except Exception as e:
            self.error.emit(f"Erro ao carregar LLM: {e}")
            self.llm_loaded.emit(False)
    
    def _process_request(self, request: Dict) -> None:
        """Processa requisição usando LLM real ou fallback."""
        prompt = request.get("prompt", "")
        request_id = request.get("id", "")
        
        self.status.emit(f"Processando: {prompt[:50]}...")
        
        if self._llm:
            try:
                # Usa LLM real
                output = self._llm(
                    prompt,
                    max_tokens=256,
                    temperature=0.1,
                    top_p=0.9,
                    stop=["\\n\\n"]
                )
                response = output["choices"][0]["text"].strip()
            except Exception as e:
                response = f"[Erro LLM: {e}]"
        else:
            # Fallback sem LLM
            response = f"[Modo offline - prompt: {prompt[:100]}...]"
        
        self.llm_response.emit(request_id, response)
    
    def queue_request(self, prompt: str, request_id: str = "") -> None:
        """Adiciona requisição à fila."""
        self._request_queue.append({
            "prompt": prompt,
            "id": request_id
        })


class RenderWorker(BaseWorker):
    """Worker de renderização SVG com integração real ao VectorEngine."""
    
    render_complete = Signal(str, str)  # job_id, output_path
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._render_queue: List[Dict] = []
    
    def run(self) -> None:
        """Loop de renderização."""
        self.status.emit("RenderWorker iniciado")
        
        while self._running:
            self.check_pause()
            
            if self._render_queue:
                job = self._render_queue.pop(0)
                self._render_job(job)
            else:
                time.sleep(0.1)
        
        self.finished_work.emit()
    
    def _render_job(self, job: Dict) -> None:
        """Renderiza um job usando VectorEngine real."""
        job_id = job.get("id", "")
        svg_path = job.get("svg_path", "")
        output_path = job.get("output_path", "")
        slots_data = job.get("slots_data", {})
        
        self.progress.emit(10, f"Renderizando {job_id}...")
        
        try:
            # 1. Carrega SVG com VectorEngine real
            from src.rendering.vector import VectorEngine
            
            engine = VectorEngine()
            self.progress.emit(20, "Carregando template SVG...")
            
            if not engine.load(svg_path):
                raise RuntimeError(f"Falha ao carregar SVG: {svg_path}")
            
            self.progress.emit(40, "Injetando dados nos slots...")
            
            # 2. Injeta dados em cada slot
            for slot_id, product_data in slots_data.items():
                engine.inject_product(int(slot_id), product_data)
            
            self.progress.emit(60, "Gerando saída...")
            
            # 3. Renderiza conforme extensão de saída
            output_ext = Path(output_path).suffix.lower()
            
            if output_ext == '.pdf':
                success = engine.render_to_pdf(output_path)
            elif output_ext in ['.png', '.jpg', '.jpeg']:
                # Usa CairoSVG se disponível, senão fallback
                try:
                    import cairosvg
                    svg_string = engine.to_string()
                    cairosvg.svg2png(bytestring=svg_string.encode('utf-8'), write_to=output_path, dpi=300)
                    success = True
                except ImportError:
                    self.error.emit("CairoSVG não instalado - usando fallback SVG")
                    engine.save(output_path.replace(output_ext, '.svg'))
                    success = True
            else:
                # Salva SVG modificado
                success = engine.save(output_path)
            
            self.progress.emit(100, "Concluído!")
            
            if success:
                self.render_complete.emit(job_id, output_path)
            else:
                self.error.emit(f"Falha na renderização de {job_id}")
                
        except Exception as e:
            self.error.emit(f"Erro de renderização: {e}")
    
    def queue_render(
        self, 
        svg_path: str, 
        output_path: str, 
        slots_data: Dict,
        job_id: str = ""
    ) -> None:
        """Adiciona job de renderização."""
        self._render_queue.append({
            "id": job_id or str(len(self._render_queue)),
            "svg_path": svg_path,
            "output_path": output_path,
            "slots_data": slots_data
        })


class ImportWorker(BaseWorker):
    """Worker de importação de Excel/dados com integração real."""
    
    row_imported = Signal(int, dict)  # row_number, data
    import_complete = Signal(int)  # total_rows
    
    def __init__(self, file_path: str = "", parent=None):
        super().__init__(parent)
        self.file_path = file_path
    
    def run(self) -> None:
        """Importa arquivo usando openpyxl ou pandas."""
        if not self.file_path:
            self.error.emit("Nenhum arquivo especificado")
            return
        
        self.status.emit(f"Importando {Path(self.file_path).name}...")
        
        try:
            file_ext = Path(self.file_path).suffix.lower()
            rows_data = []
            
            # Tenta usar openpyxl para Excel
            if file_ext in ['.xlsx', '.xls']:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
                    ws = wb.active
                    
                    # Assume primeira linha é header
                    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                        self.check_pause()
                        if not self._running:
                            break
                        
                        row_dict = dict(zip(headers, row))
                        # Normaliza campos comuns
                        normalized = {
                            "sku_origem": str(row_dict.get("SKU", row_dict.get("sku", "")) or ""),
                            "nome": str(row_dict.get("Nome", row_dict.get("nome", row_dict.get("Produto", ""))) or ""),
                            "preco": float(row_dict.get("Preço", row_dict.get("preco", row_dict.get("Valor", 0))) or 0),
                            "marca": str(row_dict.get("Marca", row_dict.get("marca", "")) or ""),
                        }
                        rows_data.append(normalized)
                    wb.close()
                except ImportError:
                    # Fallback para csv.DictReader se openpyxl não disponível
                    self.status.emit("openpyxl não disponível, tentando CSV...")
                    raise
            
            # CSV fallback
            elif file_ext == '.csv':
                import csv
                with open(self.file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        normalized = {
                            "sku_origem": row.get("SKU", row.get("sku", "")),
                            "nome": row.get("Nome", row.get("nome", row.get("Produto", ""))),
                            "preco": float(row.get("Preço", row.get("preco", row.get("Valor", 0))) or 0),
                            "marca": row.get("Marca", row.get("marca", "")),
                        }
                        rows_data.append(normalized)
            
            # Emite resultados
            total_rows = len(rows_data)
            for i, row_data in enumerate(rows_data):
                self.check_pause()
                if not self._running:
                    break
                self.progress.emit(int((i / max(total_rows, 1)) * 100), f"Linha {i+1}/{total_rows}")
                self.row_imported.emit(i, row_data)
            
            self.import_complete.emit(total_rows)
            self.status.emit(f"Importação concluída: {total_rows} linhas!")
            
        except Exception as e:
            self.error.emit(f"Erro na importação: {e}")
        
        self.finished_work.emit()


class ImageProcessWorker(BaseWorker):
    """Worker de processamento de imagem com integração real a PIL e rembg."""
    
    image_processed = Signal(str, str)  # hash_original, hash_processed
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._process_queue: List[Dict] = []
    
    def run(self) -> None:
        """Loop de processamento de imagens."""
        self.status.emit("ImageProcessWorker iniciado")
        
        while self._running:
            self.check_pause()
            
            if self._process_queue:
                task = self._process_queue.pop(0)
                self._process_image(task)
            else:
                time.sleep(0.1)
        
        self.finished_work.emit()
    
    def _process_image(self, task: Dict) -> None:
        """Processa uma imagem com PIL e rembg."""
        import hashlib
        from PIL import Image
        
        image_path = task.get("path", "")
        output_path = task.get("output_path", "")
        operations = task.get("operations", [])
        
        self.progress.emit(10, "Carregando imagem...")
        
        try:
            # Carrega imagem com PIL
            img = Image.open(image_path)
            original_hash = hashlib.md5(open(image_path, 'rb').read()).hexdigest()[:16]
            
            # Remove background se pedido
            if "remove_bg" in operations:
                self.progress.emit(40, "Removendo fundo...")
                try:
                    from rembg import remove
                    img = remove(img)
                except ImportError:
                    self.status.emit("rembg não disponível, pulando remoção de fundo")
            
            # Auto-crop se pedido
            if "autocrop" in operations:
                self.progress.emit(60, "Auto-crop inteligente...")
                bbox = img.getbbox()
                if bbox:
                    img = img.crop(bbox)
            
            # Upscale se pedido e imagem pequena
            if "upscale" in operations and min(img.size) < 500:
                self.progress.emit(80, "Upscale 2x (Lanczos)...")
                # Fallback para PIL upscale se ESRGAN não disponível
                new_size = (img.width * 2, img.height * 2)
                img = img.resize(new_size, Image.LANCZOS)
            
            self.progress.emit(90, "Salvando e calculando hash...")
            
            # Salva imagem processada
            final_path = output_path or image_path.replace('.', '_processed.')
            if img.mode == 'RGBA':
                img.save(final_path, 'PNG')
            else:
                img.save(final_path, 'JPEG', quality=95)
            
            # Calcula novo hash
            new_hash = hashlib.md5(open(final_path, 'rb').read()).hexdigest()[:16]
            
            self.progress.emit(100, "Concluído!")
            self.image_processed.emit(original_hash, new_hash)
            
        except Exception as e:
            self.error.emit(f"Erro no processamento: {e}")
    
    def queue_process(
        self, 
        image_path: str, 
        operations: List[str]
    ) -> None:
        """Adiciona tarefa de processamento."""
        self._process_queue.append({
            "path": image_path,
            "operations": operations
        })
